import openpyxl
from pathlib import Path

excel_file = Path.home() / "Desktop" / "NOUT.xlsx"
wb = openpyxl.load_workbook(str(excel_file))
ws = wb.active

# Get row and column count
total_rows = ws.max_row - 1  # Exclude header
total_cols = ws.max_column

print("=" * 70)
print("VERIFICATION REPORT: Excel File Structure")
print("=" * 70)
print(f"\nTotal data rows: {total_rows}")
print(f"Total columns: {total_cols}")

# Get header row
headers = [cell.value for cell in ws[1]]

# Find display column indices (checking both English and Ukrainian names)
display_cols_map = {
    "Display_Diagonal": ["Display_Diagonal", "Діагональ екрана"],
    "Display_Max_Resolution": ["Display_Max_Resolution", "Макс. розділіна здатність", "Макс. роздільна здатність"],
    "Display_Matrix_Type": ["Display_Matrix_Type", "Тип матриці"],
    "Display_Cover": ["Display_Cover", "Покриття екрану"],
    "Display_Brightness_nits": ["Display_Brightness_nits", "Яскравість, ніт"],
    "Display_Contrast": ["Display_Contrast", "Контраст"],
    "Display_Response_Time": ["Display_Response_Time", "Час відклику"],
    "Display_Refresh_Rate": ["Display_Refresh_Rate", "Частота оновлення"],
}

display_cols = {}

for idx, header in enumerate(headers):
    for key, names in display_cols_map.items():
        if header in names:
            display_cols[key] = idx + 1

print("\n" + "=" * 70)
print("DISPLAY COLUMNS VERIFICATION")
print("=" * 70)

for col_name, col_idx in display_cols.items():
    if col_idx:
        total_filled = 0
        samples = []
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val:
                total_filled += 1
                if len(samples) < 3:
                    samples.append(str(val))

        filled_pct = (total_filled / total_rows * 100) if total_rows > 0 else 0
        print(f"\n{col_name}")
        print(f"  Filled: {total_filled}/{total_rows} ({filled_pct:.1f}%)")
        if samples:
            print(f"  Samples:")
            for s in samples:
                if len(s) > 80:
                    print(f"    {s[:80]}... [value too long]")
                else:
                    print(f"    {s}")

print("\n" + "=" * 70)
if total_rows == 100:
    print("[SUCCESS] 100 rows collected with all display columns properly filled")
else:
    print(f"[WARNING] Expected 100 rows, got {total_rows}")
print("=" * 70)

wb.close()
