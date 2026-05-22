import re
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

BASE_URL    = "https://shop.lenovo.ua"
CATALOG_URL = "https://shop.lenovo.ua/category/notebooks"
OUTPUT_PATH = Path.home() / "Desktop" / "NOUT.xlsx"
PAGE_TIMEOUT = 30_000  # 30 сек - сайт загружается медленно
LOAD_DETAIL_PAGES = True  # Загружаем детали со страницы товара для полной информации
MAX_ITEMS = 5  # Лимит на количество товаров для сбора (100 товаров для полного анализа)

COLUMNS = [
    "Seria", "Model", "Part Number",
    "CPU_Brand", "CPU_Model", "CPU_Cores", "CPU_Threads",
    "CPU_Base_Freq_GHz", "CPU_Max_Freq_GHz", "CPU_Cache_L3_MB",
    "RAM_Type", "RAM_Freq", "RAM_Size_GB", "RAM_Slots",
    "Nakopychuvach_SSD",
    "Display_Diagonal", "Display_Max_Resolution", "Display_Matrix_Type",
    "Display_Cover", "Display_Brightness_nits", "Display_Contrast",
    "Display_Response_Time", "Display_Refresh_Rate",
    "Video_Brand", "GPU_Type", "GPU_Model", "GPU_Memory_MB",
    "OS",
    "Battery_Capacity_Wh", "Camera_MP",
    "USB_TypeA", "USB_TypeC", "HDMI", "DisplayPort", "Ports",
    "Keyboard_Waterproof", "Keyboard_Ukrainian", "Keyboard_Pointing_Device",
    "Network_3G4G", "Bluetooth", "WiFi", "LAN_Mbps",
    "Certificates", "Warranty",
    "Merezha_WiFi", "Tsina_UAH", "URL",
]

SERIES_MAP = [
    "ThinkPad", "ThinkBook", "IdeaPad", "Legion", "Yoga", "LOQ",
    "V15", "V14", "V16", "E14", "E15", "E16",
]

# Компилированные регулярные выражения для оптимизации
RE_PART_NUMBER = re.compile(r'\(([A-Z0-9-]{6,})\)')
RE_PART_NUMBER_ALT = re.compile(r'\b[A-Z0-9]{8,}\b')
RE_DIAG = re.compile(r'(\d+(?:[\.,]\d+)?)\s*(?:"|дюйм|дюйма|inch|in\b)', re.I)
RE_RESOLUTION = re.compile(r'\d{3,4}[xх]\d{3,4}|fhd|wuxga|qhd|wqxga|uhd|4k', re.I)
RE_DISPLAY_TYPE = re.compile(r'\b(ips|oled|tn|va|retina)\b', re.I)
RE_DISPLAY_COVER = re.compile(r'(антиблік|антибл|матов|глян|antiglare|gloss)', re.I)
RE_BRIGHTNESS = re.compile(r'(?:яскравість|яскравость|brightness)[,:]?\s*(?:ніт|нит|nit)?\s*(\d+)|(\d+)\s*(?:ніт|нит|nit)', re.I)
RE_CONTRAST = re.compile(r'контраст', re.I)
RE_RESPONSE_TIME = re.compile(r'(\d+)\s*мс')
RE_REFRESH_RATE = re.compile(r'(?:частота\s*оновлення|refresh\s*rate)|(?:гц.*?(?:екран|диспле|дисплей|панель|wuxga|fhd|qhd|uhd|4k|oled|ips|tn))', re.I)
RE_CPU_FREQ = re.compile(r'(\d+(?:[\.,]\d+)?)\s*ггц', re.I)
RE_CPU_CORES = re.compile(r'(\d+)\s*(?:ядер|ядра|cores?)', re.I)
RE_CPU_THREADS = re.compile(r'(\d+)\s*(?:потоків|потоки|threads?)', re.I)
RE_L3_CACHE = re.compile(r'\b(?:l3|кеш\s*l3|l3\s*кеш|cache\s*l3)\b', re.I)
RE_DDR = re.compile(r'\bDDR\d(?:\.\d)?\b', re.I)
RE_RAM_FREQ = re.compile(r'\b(\d{3,4})\s*мгц\b', re.I)
RE_RAM_SIZE = re.compile(r'\b(\d+)\s*(?:гб|gb)\b', re.I)
RE_RAM_SLOT = re.compile(r'(\d{1,2})\s*слот|[xх]\s*([12])(?=\s|[^0-9]|$)', re.I)
RE_STORAGE = re.compile(r'\b(?:ssd|hdd)\b', re.I)
RE_GPU = re.compile(r'\b(?:nvidia|geforce|rtx|gtx|radeon|intel graphics|arc|gpu|вбудован|дискретн)\b', re.I)
RE_OS = re.compile(r'\b(?:windows|linux|без ос|ubuntu|dos)\b', re.I)
RE_BATTERY = re.compile(r'(?:енергетична ємність|energy capacity|battery capacity)[,:]?\s*(?:(\d+)\s*)?(?:вт\*год|вт·год|wh|watt.?hour)?\s*(\d+)?', re.I)
RE_CAMERA = re.compile(r'\b(?:камера|web-камера|веб-камера|hd|4k)|\d+\s*(?:мп|p)\b', re.I)
RE_USB_A = re.compile(r'\b(?:usb(?:\s*type-?a|\s*a)|usb-a|type-?a)\b', re.I)
RE_USB_C = re.compile(r'\b(?:usb(?:\s*type-?c|\s*c)|usb-c|type-?c)\b', re.I)
RE_HDMI = re.compile(r'\bhdmi\b', re.I)
RE_DISPLAYPORT = re.compile(r'\b(?:displayport|dp)\b', re.I)
RE_PORTS = re.compile(r"\b(?:порт(?:и)?|роз['']єм(?:и)?|розєм(?:и)?)\b", re.I)
RE_WATERPROOF = re.compile(r'\b(?:вологозахист|водозахист|waterproof)\b', re.I)
RE_UKRAINIAN = re.compile(r'\b(?:україн|укр)\b', re.I)
RE_TOUCHPAD = re.compile(r'\b(?:тачпад|маніпулятор|трекпоінт|touchpad|trackpad)\b', re.I)
RE_3G4G = re.compile(r'\b(?:3g|4g|lte)\b', re.I)
RE_BLUETOOTH = re.compile(r'\bbt\b|bluetooth', re.I)
RE_WIFI = re.compile(r'\b(?:wi-fi|wifi|wireless)\b', re.I)
RE_LAN = re.compile(r'\b(?:lan|rj-45|ethernet)\b', re.I)
RE_WARRANTY = re.compile(r'\b(?:гарант|warranty)\b', re.I)

# Маппинг разрешений экрана на их пиксельные размеры
RESOLUTION_MAP = {
    "fhd": "1920x1080",
    "wuxga": "1920x1200",
    "qhd": "2560x1440",
    "wqxga": "2560x1600",
    "uhd": "3840x2160",
    "4k": "3840x2160",
}


def detect_series(name):
    for kw in SERIES_MAP:
        if kw.lower() in name.lower():
            return kw
    return "-"


def extract_pn(name):
    name = name.strip()
    candidates = RE_PART_NUMBER.findall(name)
    if candidates:
        return max(candidates, key=len)
    
    candidates = RE_PART_NUMBER_ALT.findall(name)
    return max(candidates, key=len) if candidates else "-"


def extract_cpu_details(cpu_model_str):
    """Извлекает детали CPU из строки модели CPU"""
    if not cpu_model_str or cpu_model_str == "-":
        return {}

    details = {}
    s = cpu_model_str.lower()

    # Извлекаем Cache L3: "(нов 24 МБ" или "cache 24mb" и т.д.
    m = re.search(r'\((?:нов|кеш\s*)?(\d+)\s*(?:мб|mb)', s)
    if m:
        details["cache"] = m.group(1) + " МБ"

    # Извлекаем частоты: "до 4.9 ГГц" или "up to 4.9 GHz" (может быть запятая 4,9)
    freq_matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:ггц|ghz)', s)
    if freq_matches:
        # Преобразуем в числа для сравнения
        freqs_numeric = []
        for f in freq_matches:
            try:
                freqs_numeric.append((float(f.replace(',', '.')), f.replace(',', '.')))
            except:
                pass

        if len(freqs_numeric) >= 2:
            freqs_numeric.sort()
            details["base_freq"] = f"{freqs_numeric[0][1]} ГГц"
            details["max_freq"] = f"{freqs_numeric[-1][1]} ГГц"
        elif len(freqs_numeric) == 1:
            # Одна частота - если есть "до/до" или "turbo/boost" это максимум, иначе базовая
            if re.search(r'\b(?:до|up to|turbo|boost|max)\b', s):
                details["max_freq"] = freqs_numeric[0][1] + " ГГц"
            else:
                details["base_freq"] = freqs_numeric[0][1] + " ГГц"

    # Попытка извлечь ядра и потоки из строки вроде "6-core, 12-thread" или "6 ядер, 12 потоков"
    cores = re.search(r'(\d+)(?:\s*-)?core|(\d+)\s*ядер', s)
    if cores:
        details["cores"] = (cores.group(1) or cores.group(2))

    threads = re.search(r'(\d+)(?:\s*-)?thread|(\d+)\s*потоків', s)
    if threads:
        details["threads"] = (threads.group(1) or threads.group(2))

    return details


def parse_specs(items):
    keys = [
        "CPU_Brand", "CPU_Model", "CPU_Cores", "CPU_Threads",
        "CPU_Base_Freq_GHz", "CPU_Max_Freq_GHz", "CPU_Cache_L3_MB",
        "RAM_Type", "RAM_Freq", "RAM_Size_GB", "RAM_Slots",
        "Nakopychuvach_SSD", "Display_Diagonal", "Display_Max_Resolution",
        "Display_Matrix_Type", "Display_Cover", "Display_Brightness_nits",
        "Display_Contrast", "Display_Response_Time", "Display_Refresh_Rate",
        "Video_Brand", "GPU_Type", "GPU_Model", "GPU_Memory_MB",
        "OS", "Battery_Capacity_Wh", "Camera_MP", "USB_TypeA", "USB_TypeC",
        "HDMI", "DisplayPort", "Ports", "Keyboard_Waterproof",
        "Keyboard_Ukrainian", "Keyboard_Pointing_Device", "Network_3G4G",
        "Bluetooth", "WiFi", "LAN_Mbps", "Certificates", "Warranty",
        "Merezha_WiFi",
    ]
    r = {k: "-" for k in keys}

    def set_if_empty(key, value):
        if r[key] == "-" and value:
            r[key] = value.rstrip(";").strip() if isinstance(value, str) else str(value)

    def set_from_detail(key, value):
        """Set value from detail page, overriding generic 'Так'/'Є' placeholders."""
        if value and (r[key] in ("-", "Так", "Є")):
            r[key] = value.rstrip(";").strip() if isinstance(value, str) else str(value)

    full_text = " ".join([str(item).lower() for item in items if item])

    # Первый проход: ищем RAM_Type (DDR тип) - это важно сделать в первую очередь
    for s in items:
        if not s:
            continue
        if RE_DDR.search(s):
            m = RE_DDR.search(s)
            if m:
                ddr_type = m.group(0)
                r["RAM_Type"] = ddr_type
                break

    for s in items:
        if not s:
            continue
        sl = s.lower()

        # Display / Экран
        if RE_DIAG.search(sl):
            m = RE_DIAG.search(sl)
            if m and r["Display_Diagonal"] == "-":
                set_if_empty("Display_Diagonal", m.group(0))
        if RE_RESOLUTION.search(sl):
            m = RE_RESOLUTION.search(sl)
            if m:
                resolution_match = m.group(0).lower()
                # Проверяем наличие и пиксельного размера и имени разрешения
                pixel_m = re.search(r'(\d{3,4})[xх](\d{3,4})', sl)
                name_m = re.search(r'\b(fhd|wuxga|qhd|wqxga|uhd|4k)\b', sl, re.I)

                if pixel_m and name_m:
                    # Оба есть - комбинируем
                    resolution_name = name_m.group(0).upper()
                    resolution_pixels = pixel_m.group(0).upper()
                    set_if_empty("Display_Max_Resolution", f"{resolution_name} ({resolution_pixels})")
                elif pixel_m:
                    # Только пиксели
                    set_if_empty("Display_Max_Resolution", pixel_m.group(0).upper())
                elif name_m:
                    # Только имя - ищем пиксели в маппинге
                    name_key = name_m.group(0).lower()
                    if name_key in RESOLUTION_MAP:
                        pixels = RESOLUTION_MAP[name_key]
                        set_if_empty("Display_Max_Resolution", f"{name_m.group(0).upper()} ({pixels})")
                    else:
                        set_if_empty("Display_Max_Resolution", name_m.group(0).upper())
                else:
                    set_if_empty("Display_Max_Resolution", resolution_match.upper())
        if RE_DISPLAY_TYPE.search(sl):
            m = RE_DISPLAY_TYPE.search(sl)
            if m:
                set_if_empty("Display_Matrix_Type", m.group(0).lower())
        if RE_DISPLAY_COVER.search(sl):
            # Извлекаем только тип покрытия
            cover_m = RE_DISPLAY_COVER.search(sl)
            if cover_m:
                cover_text = cover_m.group(0).lower()
                if "глян" in cover_text:
                    set_if_empty("Display_Cover", "Глянцеве покриття")
                elif "матов" in cover_text:
                    set_if_empty("Display_Cover", "Матове покриття")
                elif "антиблі" in cover_text or "antiglare" in cover_text:
                    set_if_empty("Display_Cover", "Антиблікове покриття")
                elif "gloss" in cover_text:
                    set_if_empty("Display_Cover", "Глянцеве покриття")
                else:
                    set_if_empty("Display_Cover", cover_text)
        if RE_BRIGHTNESS.search(sl):
            m = RE_BRIGHTNESS.search(sl)
            if m:
                # Extract the number (could be in group 1 or 2 depending on which pattern matched)
                brightness_val = m.group(1) or m.group(2)
                if brightness_val:
                    set_if_empty("Display_Brightness_nits", f"{brightness_val} nits")
        if RE_CONTRAST.search(sl):
            # Ищем значение контраста типа "1000:1" или просто пропускаем
            contrast_m = re.search(r'(?:контраст|contrast)[:\s]+(\d+(?:\.\d+)?(?:\s*:\s*\d+)?)', sl, re.I)
            if contrast_m:
                set_if_empty("Display_Contrast", contrast_m.group(1))
        if RE_RESPONSE_TIME.search(sl):
            m = RE_RESPONSE_TIME.search(sl)
            if m:
                set_if_empty("Display_Response_Time", m.group(0))
        if RE_REFRESH_RATE.search(sl):
            # Ищем конкретное значение частоты (число с "hz" или "гц")
            freq_m = re.search(r'(\d+)\s*(?:hz|гц)\b', sl, re.I)
            if freq_m:
                set_if_empty("Display_Refresh_Rate", freq_m.group(0))
            else:
                # Если есть просто число с Hz/гц, но не в стандартном формате
                freq_m = re.search(r'(?:частота\s*оновлення|refresh\s*rate)\s*[:\s]+(\d+)', sl, re.I)
                if freq_m:
                    set_if_empty("Display_Refresh_Rate", freq_m.group(1) + " Hz")

        # CPU
        if r["CPU_Brand"] == "-":
            if "intel" in sl:
                set_if_empty("CPU_Brand", "Intel")
            elif "amd" in sl or "ryzen" in sl:
                set_if_empty("CPU_Brand", "AMD")
            elif "snapdragon" in sl:
                set_if_empty("CPU_Brand", "Snapdragon")
        
        cpu_related = any(k in sl for k in ["intel", "amd", "ryzen", "snapdragon", "celeron", "pentium", "core", "xeon", "i3", "i5", "i7", "i9", "процес", "cpu"])
        if cpu_related and r["CPU_Model"] == "-":
            if "intel" in sl:
                m = re.search(r'\b(?:core i[3-9]|xeon|celeron|pentium)[^\s]*(?:\s+[a-z]?\d{3,5})?', sl)
                if m:
                    set_if_empty("CPU_Model", m.group(0).title())
                else:
                    set_if_empty("CPU_Model", s)
            elif "amd" in sl or "ryzen" in sl:
                m = re.search(r'\b(?:ryzen|athlon|a-series)[^\s]*(?:\s+[a-z]?\d{3,5})?', sl)
                if m:
                    set_if_empty("CPU_Model", m.group(0).title())
                else:
                    set_if_empty("CPU_Model", s)
            else:
                set_if_empty("CPU_Model", s)

        if RE_CPU_CORES.search(sl):
            m = RE_CPU_CORES.search(sl)
            if m:
                set_if_empty("CPU_Cores", m.group(1))
        if RE_CPU_THREADS.search(sl):
            m = RE_CPU_THREADS.search(sl)
            if m:
                set_if_empty("CPU_Threads", m.group(1))

        if cpu_related:
            # Спец. поиск для "Номіінальна частота, ГГц 1.9" format (detail page)
            # Also try "базова" for Russian
            base_freq_m = re.search(r'(?:номін|базов)\w*\s*частота[\s,ггц]*(\d+(?:[.,]\d+)?)', sl, re.I)
            max_freq_m = re.search(r'максимальна\s*частота[\s,ггц]*(\d+(?:[.,]\d+)?)', sl, re.I)

            if base_freq_m:
                set_if_empty("CPU_Base_Freq_GHz", base_freq_m.group(1) + " ГГц")
            if max_freq_m:
                set_if_empty("CPU_Max_Freq_GHz", max_freq_m.group(1) + " ГГц")

            # Стандартный поиск для "1.9 ГГц" format
            freq_matches = RE_CPU_FREQ.findall(sl)
            if len(freq_matches) >= 2:
                freqs = [float(f.replace(',', '.')) for f in freq_matches]
                low, high = min(freqs), max(freqs)
                if r["CPU_Base_Freq_GHz"] == "-":
                    set_if_empty("CPU_Base_Freq_GHz", f"{low} ГГц")
                if r["CPU_Max_Freq_GHz"] == "-":
                    set_if_empty("CPU_Max_Freq_GHz", f"{high} ГГц")
            elif len(freq_matches) == 1:
                if "до" in sl or "turbo" in sl or "boost" in sl or "max" in sl or "up to" in sl:
                    if r["CPU_Max_Freq_GHz"] == "-":
                        set_if_empty("CPU_Max_Freq_GHz", freq_matches[0] + " ГГц")
                else:
                    if r["CPU_Base_Freq_GHz"] == "-":
                        set_if_empty("CPU_Base_Freq_GHz", freq_matches[0] + " ГГц")

        if RE_L3_CACHE.search(sl):
            # Ищем "L3 ... 24 МБ" pattern или просто "L3 24" (может быть написано как ĭ L3 24)
            m = re.search(r'l3[\s,]*(\d+)', sl, re.I)
            if m:
                set_if_empty("CPU_Cache_L3_MB", m.group(1) + " МБ")

        # RAM Memory
        if RE_RAM_FREQ.search(sl):
            m = RE_RAM_FREQ.search(sl)
            if m:
                set_if_empty("RAM_Freq", m.group(0) + " МГц")
        
        if r["RAM_Size_GB"] == "-" and "ssd" not in sl and "hdd" not in sl:
            m = RE_RAM_SIZE.search(sl)
            if m:
                value = m.group(0).upper()
                set_if_empty("RAM_Size_GB", value)

        # Ищем слоты памяти только если строка содержит признаки памяти (ОЗУ, DDR, ГБ) и слово "слот"
        if "слот" in sl and any(k in sl for k in ["озу", "ddr", "пам", "гб"]):
            m = RE_RAM_SLOT.search(sl)
            if m:
                slot_num = m.group(1) or m.group(2)
                if slot_num:
                    set_if_empty("RAM_Slots", slot_num + " слотів")

        # Storage
        if RE_STORAGE.search(sl):
            set_if_empty("Nakopychuvach_SSD", s)

        # GPU
        if RE_GPU.search(sl):
            sl_lower = sl.lower()

            # Extract GPU Memory - look for "обсяг gpu" or just numbers after GPU keywords
            if r["GPU_Memory_MB"] == "-":
                # First try: memory value with units (GB/ГБ/МБ)
                m = re.search(r'\b(\d+)\s*(?:гб|gb)\b', sl, re.I)
                if m:
                    set_if_empty("GPU_Memory_MB", m.group(1) + " GB")
                else:
                    # Second try: just memory number if it follows GPU memory label
                    if "обсяг gpu" in sl_lower or "gpu memory" in sl_lower:
                        m = re.search(r'\b(\d+)\b', sl)
                        if m:
                            set_if_empty("GPU_Memory_MB", m.group(1) + " GB")

            # Extract GPU Brand
            if r["Video_Brand"] == "-":
                if "nvidia" in sl_lower or "geforce" in sl_lower or "rtx" in sl_lower or "gtx" in sl_lower:
                    set_if_empty("Video_Brand", "NVIDIA")
                elif "amd" in sl_lower or "radeon" in sl_lower:
                    set_if_empty("Video_Brand", "AMD")
                elif "intel" in sl_lower or "arc" in sl_lower:
                    set_if_empty("Video_Brand", "Intel")
                elif "qualcomm" in sl_lower or "adreno" in sl_lower:
                    set_if_empty("Video_Brand", "Qualcomm")

            # Extract GPU Model
            if r["GPU_Model"] == "-":
                # Try discrete GPU patterns: RTX/GTX XXXX, Radeon RX XXXX
                m = re.search(r'\b(?:rtx|gtx|radeon\s+rx)\s*(\d{3,4})\b', sl, re.I)
                if m:
                    prefix = re.search(r'(rtx|gtx|radeon\s+rx)', sl, re.I)
                    model_str = f"{prefix.group(0).upper()} {m.group(1)}"
                    set_if_empty("GPU_Model", model_str)
                else:
                    # Try integrated GPU patterns: Radeon 780M, Arc, Iris, Intel Graphics, Qualcomm Adreno
                    m = re.search(r'\b(?:radeon\s+\d{3}m|arc\s+\w+|iris\s+\w+|intel\s+graphics|(?:intel\s+)?(?:uhd|hd|iris)\s+graphics|(?:qualcomm\s+)?adreno\s+gpu)\b', sl, re.I)
                    if m:
                        model_text = m.group(0).strip()
                        # Capitalize properly for Intel Graphics
                        if 'graphics' in model_text.lower():
                            # Convert "intel uhd graphics" to "Intel UHD Graphics"
                            parts = model_text.split()
                            model_text = ' '.join([p.upper() if p.upper() in ['UHD', 'HD', 'XE'] else p.title() for p in parts])
                        set_if_empty("GPU_Model", model_text)
                    else:
                        # Fallback: grab text after brand keyword, limited length
                        brand_m = re.search(r'(?:nvidia|geforce|amd|radeon|intel|arc|qualcomm|adreno)', sl, re.I)
                        if brand_m:
                            start = brand_m.start()
                            gpu_text = sl[start:].strip()
                            # Clean up extra specs (memory, etc)
                            gpu_text = re.sub(r',.*', '', gpu_text).strip()
                            gpu_text = re.sub(r'\d+\s*(?:гб|gb|мб|mb)', '', gpu_text, flags=re.I).strip()
                            # Remove part numbers in parentheses (e.g., "(83Ly00Tgra)")
                            gpu_text = re.sub(r'\s*\([^)]*\)', '', gpu_text).strip()
                            if len(gpu_text) > 3 and len(gpu_text) < 60:
                                set_if_empty("GPU_Model", gpu_text.title())

            # Determine GPU Type (integrated/discrete) based on extracted data
            if r["GPU_Type"] == "-":
                # Check if discrete GPU
                if r["Video_Brand"] == "NVIDIA" or "rtx" in sl_lower or "gtx" in sl_lower or "radeon rx" in sl_lower:
                    set_if_empty("GPU_Type", "дискретна")
                # Check if integrated GPU
                elif r["Video_Brand"] == "Intel" or r["Video_Brand"] == "Qualcomm" or "radeon 780m" in sl_lower or "iris" in sl_lower or "arc" in sl_lower or "graphics" in sl_lower or "adreno" in sl_lower:
                    set_if_empty("GPU_Type", "інтегрована")
                # AMD APU check
                elif r["Video_Brand"] == "AMD" and "radeon 680m" in sl_lower or "radeon 780m" in sl_lower:
                    set_if_empty("GPU_Type", "інтегрована")
                elif r["Video_Brand"] == "AMD":
                    set_if_empty("GPU_Type", "дискретна")

        # OS
        if RE_OS.search(sl):
            m = RE_OS.search(sl)
            if m:
                set_if_empty("OS", m.group(0).title())

        # Battery Capacity
        if RE_BATTERY.search(sl):
            m = RE_BATTERY.search(sl)
            if m:
                # Захватываем число, которое может быть в группе 1 или 2 (до или после единиц измерения)
                val = m.group(1) or m.group(2)
                if val:
                    set_if_empty("Battery_Capacity_Wh", val + " Wh")

        # Camera
        if RE_CAMERA.search(sl):
            # Извлекаем информацию о WEB-камере: текст между "WEB-камера" и "3D камера"
            # Примеры: "WEB-камера, Мп HD 720p with E-shutter 3D камера Немає"
            m = re.search(r'WEB-?[Кк]амера[,\s]*([^3]*?)\s*(?:3D|IR)\s*[Кк]амера', s, re.I)
            if m:
                camera_text = m.group(1).strip()
                # Убираем "Мп" если оно есть в начале
                camera_text = re.sub(r'^[Мм][Пп]\s*', '', camera_text)
                if camera_text and camera_text != "Немає":
                    set_if_empty("Camera_MP", camera_text)

        # Ports & Keyboard — split on ‖ separator and match by label prefix
        if "‖" in s:
            usb_a_parts, usb_c_parts, dp_parts = [], [], []
            for entry in s.split("‖"):
                e = entry.strip()
                el = e.lower()
                if re.match(r'usb\s*type-?a\s', el):
                    val = re.sub(r'^USB\s*Type-?A\s*', '', e, flags=re.I).strip()
                    if val:
                        usb_a_parts.append(val)
                elif re.match(r'usb[\s-]*type-?c\s', el):
                    val = re.sub(r'^usb[\s-]*type-?c\s*', '', e, flags=re.I).strip()
                    if val:
                        usb_c_parts.append(val)
                elif re.match(r'hdmi[,\s]', el):
                    val = re.sub(r'^hdmi[,\s]*(?:шт\.?\s+)?', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("HDMI", val)
                elif re.match(r'displayport[,\s]', el):
                    val = re.sub(r'^displayport[,\s]*(?:шт\.?\s+)?', '', e, flags=re.I).strip()
                    if val:
                        dp_parts.append(val)
                elif re.match(r'порти\s', el):
                    val = re.sub(r'^порти\s*', '', e, flags=re.I).strip()
                    if val and val.lower() != "немає":
                        set_if_empty("Ports", val)
                elif re.match(r'клавіатура\s*\(вологозахист\)', el):
                    val = re.sub(r'^клавіатура\s*\(вологозахист\)\s*', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("Keyboard_Waterproof", val)
                elif re.match(r'українська\s*мова\s', el):
                    val = re.sub(r'^українська\s*мова\s*', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("Keyboard_Ukrainian", val)
                elif re.match(r'маніпулятори\s', el):
                    val = re.sub(r'^маніпулятори\s*', '', e, flags=re.I).strip()
                    if val and val.lower() != "немає":
                        set_if_empty("Keyboard_Pointing_Device", val)
                elif re.match(r'3g/4g\s', el):
                    val = re.sub(r'^3g/4g\s*', '', e, flags=re.I).strip()
                    if val:
                        set_from_detail("Network_3G4G", val)
                elif re.match(r'bluetooth\s', el):
                    val = re.sub(r'^bluetooth\s*', '', e, flags=re.I).strip()
                    if val and val.lower() != "немає":
                        set_from_detail("Bluetooth", val)
                elif re.match(r'wi-fi\s', el):
                    val = re.sub(r'^wi-fi\s*', '', e, flags=re.I).strip()
                    if val and val.lower() != "немає":
                        set_from_detail("WiFi", val)
                        set_from_detail("Merezha_WiFi", val)
                elif re.match(r'lan\s*rj-45', el):
                    val = re.sub(r'^lan\s*rj-45[,\s]*(?:мбіт/с\s*)?', '', e, flags=re.I).strip()
                    if val and val.lower() != "немає":
                        set_from_detail("LAN_Mbps", val)
                elif re.match(r'сертифікати\s', el):
                    val = re.sub(r'^сертифікати\s*', '', e, flags=re.I).strip()
                    if val:
                        set_from_detail("Certificates", val)
                elif re.match(r'термін\s*базово', el):
                    m = re.search(r'(\d+\s*(?:рік|роки|років|місяц\w*))', e, re.I)
                    if m:
                        set_from_detail("Warranty", m.group(1).strip())
            if usb_a_parts:
                set_if_empty("USB_TypeA", " / ".join(usb_a_parts))
            if usb_c_parts:
                set_if_empty("USB_TypeC", " / ".join(usb_c_parts))
            if dp_parts:
                set_if_empty("DisplayPort", " / ".join(dp_parts))
        else:
            # Fallback for items without ‖ separator (catalog card items)
            if RE_USB_A.search(sl):
                set_if_empty("USB_TypeA", s.strip())
            if RE_USB_C.search(sl):
                set_if_empty("USB_TypeC", s.strip())
            if RE_HDMI.search(sl):
                set_if_empty("HDMI", s.strip())
            if RE_DISPLAYPORT.search(sl):
                set_if_empty("DisplayPort", s.strip())
            if RE_PORTS.search(sl):
                set_if_empty("Ports", s.strip())
            if RE_WATERPROOF.search(sl):
                set_if_empty("Keyboard_Waterproof", "Так")
            if RE_UKRAINIAN.search(sl):
                set_if_empty("Keyboard_Ukrainian", "Так")
            if RE_TOUCHPAD.search(sl):
                set_if_empty("Keyboard_Pointing_Device", "Так")
            if RE_3G4G.search(sl):
                set_if_empty("Network_3G4G", "Так")
            if RE_BLUETOOTH.search(sl):
                set_if_empty("Bluetooth", "Так")
            if RE_WIFI.search(sl):
                set_if_empty("WiFi", "Так")
                set_if_empty("Merezha_WiFi", "Так")
            if RE_LAN.search(sl):
                m = re.search(r'(\d{3,4})\s*мбіт', sl)
                if m:
                    set_if_empty("LAN_Mbps", m.group(0))
                else:
                    set_if_empty("LAN_Mbps", "Так")

        # Warranty
        if RE_WARRANTY.search(sl):
            m = re.search(r'(\d+)\s*(?:місяц|рік|year|month)', sl, re.I)
            if m:
                set_if_empty("Warranty", m.group(0))
            else:
                set_if_empty("Warranty", "Так")

    # Извлекаем дополнительные CPU детали из модели CPU и полного текста
    if r["CPU_Model"] != "-":
        cpu_details = extract_cpu_details(r["CPU_Model"])
        if "cache" in cpu_details:
            set_if_empty("CPU_Cache_L3_MB", cpu_details["cache"])
        if "base_freq" in cpu_details:
            set_if_empty("CPU_Base_Freq_GHz", cpu_details["base_freq"])
        if "max_freq" in cpu_details:
            set_if_empty("CPU_Max_Freq_GHz", cpu_details["max_freq"])
        if "cores" in cpu_details:
            set_if_empty("CPU_Cores", cpu_details["cores"])
        if "threads" in cpu_details:
            set_if_empty("CPU_Threads", cpu_details["threads"])

    # Ищем ядра и потоки во всем тексте спецификаций если еще не нашли
    if r["CPU_Cores"] == "-" or r["CPU_Threads"] == "-":
        # Ищем "Кількість ядер" или "Кількість потоків" прямо перед числом
        cores_m = re.search(r'кількість ядер\s+(\d+)', full_text, re.I)
        threads_m = re.search(r'кількість потоків\s+(\d+)', full_text, re.I)

        if cores_m:
            set_if_empty("CPU_Cores", cores_m.group(1))
        if threads_m:
            set_if_empty("CPU_Threads", threads_m.group(1))

        # Fallback: поиск просто "ядер" и "потоків" со словом
        if not cores_m:
            cores_m = re.search(r'(\d+)\s*(?:ядер|ядра|cores?)', full_text, re.I)
            if cores_m:
                set_if_empty("CPU_Cores", cores_m.group(1))
        if not threads_m:
            threads_m = re.search(r'(\d+)\s*(?:потоків|потоки|threads?)', full_text, re.I)
            if threads_m:
                set_if_empty("CPU_Threads", threads_m.group(1))

    # Установка значений по умолчанию для часто встречаемых параметров
    if r["Bluetooth"] == "-" and "lenovo" in full_text:
        set_if_empty("Bluetooth", "Так")
    if r["WiFi"] == "-" and "lenovo" in full_text:
        set_if_empty("WiFi", "Так")
        set_if_empty("Merezha_WiFi", "Так")

    return r


def fetch_detail_page_text(pg, url):
    try:
        pg.goto(url, wait_until="load", timeout=PAGE_TIMEOUT)

        # Извлекаем спецификации из HTML напрямую
        html = pg.content()
        soup = BeautifulSoup(html, "lxml")

        specs_text_parts = []

        # Ищем все label divs и находим их соответствующие value divs (siblings)
        labels = soup.select("div.product-full-specs-label")
        for label_div in labels:
            label_span = label_div.select_one("span")
            if label_span:
                label_text = label_span.get_text(strip=True)

                # Ищем следующий sibling div с классом product-full-specs-value
                current = label_div.next_sibling
                while current:
                    if hasattr(current, 'name') and current.name == 'div' and 'product-full-specs-value' in current.get('class', []):
                        value_text = current.get_text(separator=" ", strip=True)
                        if label_text and value_text:
                            # Формируем строку как "Label Value" для совместимости с regex
                            specs_text_parts.append(f"{label_text} {value_text}")
                        break
                    current = current.next_sibling

        if specs_text_parts:
            # Use ‖ as separator so per-field regex can stop cleanly at the next spec
            specs_text = " ‖ ".join(specs_text_parts)
            return specs_text

        return ""
    except Exception:
        return ""


def parse_cards_from_html(html, detail_page=None):
    soup = BeautifulSoup(html, "lxml")
    records = []
    for card in soup.select("article.product-card"):
        link_el = card.select_one("a.product-card-link")
        if not link_el:
            continue
        name  = link_el.get_text(strip=True)
        href  = link_el.get("href", "")
        url   = BASE_URL + href if href.startswith("/") else href
        price_el = card.select_one("p.price")
        price = re.sub(r"[^\d\s]", "", price_el.get_text()).strip() if price_el else "-"
        specs_items = [name] + [li.get_text(strip=True)
                                for li in card.select("li.product-specs-item")]
        if detail_page is not None and LOAD_DETAIL_PAGES:
            # Доповнюємо дані повною сторінкою товару для CPU-специфікацій
            detail_text = fetch_detail_page_text(detail_page, url)
            if detail_text:
                specs_items.append(detail_text)
        specs = parse_specs(specs_items)
        records.append({
            "Seria":             detect_series(name),
            "Model":             name,
            "Part Number":       extract_pn(name),
            "CPU_Brand":         specs["CPU_Brand"],
            "CPU_Model":         specs["CPU_Model"],
            "CPU_Cores":         specs["CPU_Cores"],
            "CPU_Threads":       specs["CPU_Threads"],
            "CPU_Base_Freq_GHz": specs["CPU_Base_Freq_GHz"],
            "CPU_Max_Freq_GHz":  specs["CPU_Max_Freq_GHz"],
            "CPU_Cache_L3_MB":   specs["CPU_Cache_L3_MB"],
            "RAM_Type":          specs["RAM_Type"],
            "RAM_Freq":          specs["RAM_Freq"],
            "RAM_Size_GB":       specs["RAM_Size_GB"],
            "RAM_Slots":         specs["RAM_Slots"],
            "Nakopychuvach_SSD": specs["Nakopychuvach_SSD"],
            "Display_Diagonal":  specs["Display_Diagonal"],
            "Display_Max_Resolution": specs["Display_Max_Resolution"],
            "Display_Matrix_Type": specs["Display_Matrix_Type"],
            "Display_Cover":     specs["Display_Cover"],
            "Display_Brightness_nits": specs["Display_Brightness_nits"],
            "Display_Contrast":  specs["Display_Contrast"],
            "Display_Response_Time": specs["Display_Response_Time"],
            "Display_Refresh_Rate": specs["Display_Refresh_Rate"],
            "Video_Brand":       specs["Video_Brand"],
            "GPU_Type":          specs["GPU_Type"],
            "GPU_Model":         specs["GPU_Model"],
            "GPU_Memory_MB":     specs["GPU_Memory_MB"],
            "OS":                specs["OS"],
            "Battery_Capacity_Wh": specs["Battery_Capacity_Wh"],
            "Camera_MP":         specs["Camera_MP"],
            "USB_TypeA":         specs["USB_TypeA"],
            "USB_TypeC":         specs["USB_TypeC"],
            "HDMI":              specs["HDMI"],
            "DisplayPort":       specs["DisplayPort"],
            "Ports":             specs["Ports"],
            "Keyboard_Waterproof":   specs["Keyboard_Waterproof"],
            "Keyboard_Ukrainian":     specs["Keyboard_Ukrainian"],
            "Keyboard_Pointing_Device": specs["Keyboard_Pointing_Device"],
            "Network_3G4G":      specs["Network_3G4G"],
            "Bluetooth":         specs["Bluetooth"],
            "WiFi":              specs["WiFi"],
            "LAN_Mbps":          specs["LAN_Mbps"],
            "Certificates":      specs["Certificates"],
            "Warranty":          specs["Warranty"],
            "Merezha_WiFi":      specs["Merezha_WiFi"],
            "Tsina_UAH":         price,
            "URL":               url,
        })
    return records


def load_page_and_wait(pg, url, page_num):
    """
    Завантажує сторінку каталогу і чекає появи карток.
    Повертає True якщо картки знайдені.
    """
    attempt = 0
    while attempt < 2:
        try:
            pg.goto(url, wait_until="load", timeout=PAGE_TIMEOUT)
            break
        except Exception as err:
            attempt += 1
            if attempt >= 2:
                print(f"Помилка завантаження: {err}")
                return False
            time.sleep(0.5)

    # Чекаємо поки React відрендерить картки (до 10 сек)
    deadline = time.time() + 10
    cards_found = False
    last_count = 0

    while time.time() < deadline:
        try:
            # Используем evaluate для получения количества карточек напрямую
            count = pg.evaluate("document.querySelectorAll('article.product-card').length")
            if count > 0:
                if count == last_count:
                    # Количество не меняется - значит загрузка завершена
                    cards_found = True
                    break
                last_count = count
                print(f"  Найдено {count} карточек...")
        except Exception as e:
            print(f"  Ошибка при проверке: {e}")
        time.sleep(0.2)

    # Остання спроба — просто беремо що є
    cards = pg.query_selector_all("article.product-card")
    found = len(cards) > 0

    if not found:
        print(f"  Карточки не найдены! Сохраняю HTML для отладки...")
        # Сохраняем HTML для анализа
        html = pg.content()
        with open("debug_page.html", "w", encoding="utf-8") as f:
            f.write(html)
        print(f"  HTML сохранен в debug_page.html")

        # Проверяем другие селекторы
        selectors = [
            "div.product-card",
            "div[class*='product']",
            "article",
            "[class*='card']"
        ]
        for sel in selectors:
            try:
                count = len(pg.query_selector_all(sel))
                if count > 0:
                    print(f"  Найдено {count} элементов по селектору: {sel}")
            except:
                pass

    return found


def format_xlsx(path, total):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Lenovo Notebooks"

    UA = {
        "Seria":"Серія", "Model":"Модель", "Part Number":"Part Number",
        "CPU_Brand":"Бренд CPU", "CPU_Model":"Модель CPU", "CPU_Cores":"Кількість ядер",
        "CPU_Threads":"Кількість потоків", "CPU_Base_Freq_GHz":"Номінальна частота, ГГц",
        "CPU_Max_Freq_GHz":"Максимальна частота, ГГц", "CPU_Cache_L3_MB":"Кеш L3, МБ",
        "RAM_Type":"Тип ОЗП", "RAM_Freq":"Частота ОЗП",
        "RAM_Size_GB":"Обсяг ОЗП, ГБ", "RAM_Slots":"Кількість слотів",
        "Nakopychuvach_SSD":"M.2 SSD, ГБ",
        "Display_Diagonal":"Діагональ екрана", "Display_Max_Resolution":"Макс. роздільна здатність",
        "Display_Matrix_Type":"Тип матриці", "Display_Cover":"Покриття екрану",
        "Display_Brightness_nits":"Яскравість, ніт", "Display_Contrast":"Контраст",
        "Display_Response_Time":"Час реагування", "Display_Refresh_Rate":"Частота оновлення",
        "Video_Brand":"Видео_Бренд", "GPU_Type":"Тип_GPU", "GPU_Model":"Модель GPU",
        "GPU_Memory_MB":"Обсяг GPU, МБ", "OS":"ОС", "Battery_Capacity_Wh":"Енергетична ємність, Вт*год",
        "Camera_MP":"WEB-камера, Мп", "USB_TypeA":"USB Type-A", "USB_TypeC":"USB Type-C",
        "HDMI":"HDMI, шт.", "DisplayPort":"DisplayPort, шт.", "Ports":"Порти",
        "Keyboard_Waterproof":"Клавіатура (вологозахист)", "Keyboard_Ukrainian":"Українська мова",
        "Keyboard_Pointing_Device":"Маніпулятори", "Network_3G4G":"3G/4G",
        "Bluetooth":"Bluetooth", "WiFi":"Wi-Fi", "LAN_Mbps":"LAN RJ-45, Мбіт/с",
        "Certificates":"Сертифікати", "Warranty":"Термін базової гарантії від виробника",
        "Merezha_WiFi":"Мережа / Wi-Fi", "Tsina_UAH":"Ціна (UAH)", "URL":"URL",
    }
    for i, k in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=i).value = UA.get(k, k)

    HFILL = PatternFill("solid", fgColor="1F3864")
    AFILL = PatternFill("solid", fgColor="EEF2FF")
    t = Side(style="thin", color="BBBBBB")
    brd = Border(left=t, right=t, top=t, bottom=t)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
    ws.row_dimensions[1].height = 32

    for r in range(2, ws.max_row + 1):
        for cell in ws[r]:
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = AFILL if r % 2 == 0 else PatternFill()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = brd
        ws.row_dimensions[r].height = 45

    for i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=len(COLUMNS))
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    info = wb.create_sheet("Info")
    info["A1"], info["B1"] = "Sformovano:", datetime.now().strftime("%d.%m.%Y %H:%M")
    info["A2"], info["B2"] = "Dzherelo:",   CATALOG_URL
    info["A3"], info["B3"] = "Modelej:",    total
    for c in ["A1","A2","A3"]:
        info[c].font = Font(bold=True, name="Arial")
    wb.save(path)


def main():
    print("=" * 60)
    print("  Parser noutkbukiv shop.lenovo.ua")
    print("=" * 60)

    all_records = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"]
        )
        ctx = browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            locale="uk-UA",
        )
        pg = ctx.new_page()

        # Сторінка 1 — визначаємо кількість сторінок
        print("\nZavantazhuiemo storinkу 1...")
        ok = load_page_and_wait(pg, CATALOG_URL, 1)
        if not ok:
            print("POMYLKA: kartky ne znajdeno.")
            browser.close()
            return

        page_btns = pg.query_selector_all("button.MuiPaginationItem-page")
        nums = [int(b.inner_text().strip())
                for b in page_btns if b.inner_text().strip().isdigit()]
        total_pages = max(nums) if nums else 43
        print("Storinok: " + str(total_pages) + "  (~" + str(total_pages * 24) + " tovariv)")
        print()

        # Визначаємо URL-схему пагінації — пробуємо ?page=2
        # Якщо не спрацює — буде fallback

        # Используем только '?page={page}' — это работает всегда
        test_url_pattern = CATALOG_URL + "?page={page}"

        # Повертаємось на сторінку 1 і парсимо
        load_page_and_wait(pg, CATALOG_URL, 1)
        recs = parse_cards_from_html(pg.content(), pg)
        all_records.extend(recs)
        print("  [1/" + str(total_pages) + "] zibrano " + str(len(recs)) +
              " (vsogo: " + str(len(all_records)) + ")")

        # Обходимо сторінки 2..N
        for page_num in range(2, total_pages + 1):
            # Зупиняємо коли зібрано достатньо для тестування
            if len(all_records) >= MAX_ITEMS:
                print(f"\nДосягнуто ліміту {MAX_ITEMS} товарів.")
                break

            print("  [" + str(page_num) + "/" + str(total_pages) + "] ...", end=" ", flush=True)

            success = False
            retry_count = 0

            while not success and retry_count < 2:
                # Спосіб 1: URL-навігація
                url = test_url_pattern.replace("{page}", str(page_num))
                ok = load_page_and_wait(pg, url, page_num)
                if ok:
                    recs = parse_cards_from_html(pg.content(), pg)
                    if len(recs) > 0:
                        all_records.extend(recs)
                        print("zibrano " + str(len(recs)) + " (vsogo: " + str(len(all_records)) + ")")
                        success = True
                        break

                # Повтор при неудаче
                if not success:
                    retry_count += 1
                    if retry_count < 2:
                        print("\n  ⚠ Retry...", end=" ", flush=True)

            if not success:
                print("0 - skip")

        browser.close()

    # Збереження
    print()
    # Обмежуємо до MAX_ITEMS для тестування
    all_records = all_records[:MAX_ITEMS]
    print("Zberihaemo Excel (" + str(len(all_records)) + " modelej)...")
    df = pd.DataFrame(all_records, columns=COLUMNS)
    before = len(df)
    valid_mask = df["Part Number"] != "-"
    valid_df = df[valid_mask].drop_duplicates(subset=["Part Number"], keep="first")
    invalid_df = df[~valid_mask]
    df = pd.concat([valid_df, invalid_df], ignore_index=True)
    if len(df) < before:
        print("Dublikativ vylucheno: " + str(before - len(df)))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False, sheet_name="Lenovo Notebooks")
    print("Formatuvannia...")
    format_xlsx(OUTPUT_PATH, len(df))

    print()
    print("=" * 60)
    print("GOTOVO!  " + str(OUTPUT_PATH))
    print("Modelej u tablyci: " + str(len(df)))
    print("=" * 60)


if __name__ == "__main__":
    main()
