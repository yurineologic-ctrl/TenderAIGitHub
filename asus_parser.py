import re
import sys
import time
from datetime import datetime
from pathlib import Path

# Force UTF-8 output so Cyrillic and special chars print correctly on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

BASE_URL    = "https://www.asus.com"
CATALOG_URL = "https://www.asus.com/ua-ua/store/laptops/"
OUTPUT_PATH = Path.home() / "Desktop" / "NOUT_Asus.xlsx"
PAGE_TIMEOUT        = 45_000   # каталог
DETAIL_PAGE_TIMEOUT = 45_000   # детальні сторінки товарів
LOAD_DETAIL_PAGES   = True
MAX_ITEMS = 9999

COLUMNS = [
    "Seria", "Model", "Part Number",
    "CPU_Brand", "CPU_Model", "CPU_Cores", "CPU_Threads",
    "CPU_Base_Freq_GHz", "CPU_Max_Freq_GHz", "CPU_Cache_L3_MB",
    "RAM_Type", "RAM_Freq", "RAM_Size_GB", "RAM_Slots",
    "Nakopychuvach_SSD",
    "Display_Diagonal", "Display_Max_Resolution", "Display_Matrix_Type",
    "Display_Cover", "Display_Brightness_nits", "Display_Refresh_Rate",
    "Video_Brand", "GPU_Type", "GPU_Model", "GPU_Memory_MB",
    "OS",
    "Battery_Capacity_Wh", "Camera_MP",
    "USB_TypeA", "USB_TypeC", "HDMI",
    "Keyboard_Waterproof", "Keyboard_Ukrainian", "Keyboard_Pointing_Device",
    "Network_3G4G", "Bluetooth", "WiFi", "LAN_Mbps",
    "Certificates", "Warranty",
    "Merezha_WiFi", "Tsina_UAH", "URL",
]

# CPU spec lookup: model_substring -> (cores, threads, base_GHz, max_GHz, cache_MB)
# Specific models first; generic fallbacks at the end (iterated in insertion order)
CPU_SPECS = {
    # Intel Core Ultra 200H series (Arrow Lake-H)
    "285H":  (20, 20, "2.0", "5.4", "24"),
    "265H":  (16, 16, "2.0", "5.3", "24"),
    "255H":  (16, 16, "2.0", "5.1", "24"),
    "245H":  (14, 14, "2.0", "5.1", "24"),
    "235H":  (14, 14, "1.9", "5.0", "24"),
    "225H":  (14, 14, "1.7", "4.9", "18"),
    # Intel Core Ultra 200HX series (Arrow Lake-HX)
    "295HX": (24, 24, "2.5", "5.6", "36"),
    "290HX": (24, 24, "2.5", "5.4", "36"),
    "285HX": (24, 24, "2.4", "5.5", "36"),
    "275HX": (20, 20, "2.2", "5.4", "30"),
    "265HX": (20, 20, "2.0", "5.3", "30"),
    "255HX": (20, 20, "1.8", "5.1", "30"),
    # Intel Core Ultra 300H/HX series
    "385H":  (24, 24, "2.0", "5.5", "36"),
    "375H":  (20, 20, "2.0", "5.4", "36"),
    "365H":  (18, 18, "2.0", "5.2", "30"),
    "386H":  (24, 24, "2.0", "5.5", "36"),
    # Intel Core Ultra 100H series (Meteor Lake)
    "185H":  (16, 22, "2.3", "5.1", "24"),
    "175H":  (16, 22, "2.3", "4.8", "24"),
    "165H":  (16, 22, "1.4", "5.0", "24"),
    "155H":  (16, 22, "1.4", "4.8", "24"),
    "165U":  (12, 16, "0.9", "4.9", "12"),
    # Intel Core 200H (Arrow Lake non-Ultra)
    "240H":  (10, 16, "2.5", "5.2", "24"),
    "220H":  (10, 16, "2.4", "4.9", "20"),
    # Snapdragon X Elite / Plus / X2
    "X2 Elite":  (12, 12, "4.0", "4.6", "42"),
    "X Elite":   (12, 12, "3.8", "4.3", "42"),
    "X Plus":    (10, 10, "3.4", "4.0", "42"),
    # AMD Ryzen 7000 series (HS/HX)
    "7940HS": (8, 16, "4.0", "5.2", "16"),
    "7945HX": (16, 32, "2.5", "5.4", "64"),
    "7735HS": (8, 16, "3.2", "4.75", "16"),
    "7535HS": (6, 12, "3.3", "4.55", "16"),
    "7435HS": (8, 16, "3.1", "4.55", "20"),
    "7445HS": (6, 12, "3.2", "4.7",  "22"),
    "7745HX": (8, 16, "3.6", "5.1",  "32"),
    # AMD Ryzen AI Max / AI series (specific first)
    "AI Max 395": (16, 32, "3.8", "5.0", "64"),
    "AI Max 390": (12, 24, "3.6", "5.0", "64"),
    "AI Max":     (16, 32, "3.8", "5.0", "64"),  # generic AI Max fallback
    "AI 9 HX":    (12, 24, "3.8", "5.1", "24"),
    "AI 7":       (8,  16, "2.0", "5.0", "16"),
    "AI 5":       (8,  16, "2.0", "4.8", "16"),
    # Generic Intel Core Ultra tier fallbacks (no specific SKU in model string)
    "Ultra 9": (20, 20, "2.2", "5.4", "30"),
    "Ultra 7": (16, 16, "1.9", "5.1", "24"),
    "Ultra 5": (14, 14, "1.7", "4.9", "18"),
    # Intel Core Ultra X9/X7 (ExpertBook / ZenBook Series 3)
    "Ultra X9": (20, 20, "2.0", "5.4", "24"),
    "Ultra X7": (16, 16, "1.7", "5.1", "24"),
}

SERIES_MAP = [
    "ROG", "TUF", "ZenBook", "VivoBook", "Vivobook", "ProArt",
    "ExpertBook", "Chromebook", "StudioBook", "OLED", "Flow",
    "Zenbook", "Expertbook",
]

# Маппінг 2-символьного префіксу артикула -> серія
ASUS_PN_PREFIX_SERIES = {
    "GU": "ROG", "GX": "ROG", "G5": "ROG", "G6": "ROG",
    "G7": "ROG", "G8": "ROG", "G9": "ROG",
    "FA": "TUF", "FX": "TUF",
    "UX": "ZenBook", "UM": "ZenBook", "UP": "ZenBook",
    "BU": "ExpertBook", "B1": "ExpertBook", "B9": "ExpertBook",
    "S3": "VivoBook", "S5": "VivoBook",
    "K5": "VivoBook", "K4": "VivoBook", "K3": "VivoBook",
    "X1": "VivoBook", "X4": "VivoBook", "X5": "VivoBook",
    "M1": "VivoBook", "M3": "VivoBook", "M5": "VivoBook",
    "PA": "ProArt",
}

# Суфікси-маркетинг, які треба прибирати з назви моделі
_NAME_SUFFIXES = re.compile(
    r'\s*[;,]\s*(?:Copilot\+?\s*PC|AI\s*PC|Gaming|New\s*\d{4})\s*$', re.I
)

# ── Compiled regex patterns (same as Lenovo parser) ──────────────────────────
RE_PART_NUMBER     = re.compile(r'\(([A-Z0-9_-]{6,})\)')
RE_PART_NUMBER_ALT = re.compile(r'\b[A-Z0-9]{8,}\b')
RE_DIAG            = re.compile(r'(\d+(?:[\.,]\d+)?)\s*(?:"|дюйм|дюйма|inch|in\b)', re.I)
RE_RESOLUTION      = re.compile(r'\d{3,4}[xх]\d{3,4}|fhd|wuxga|qhd|wqxga|uhd|4k', re.I)
RE_DISPLAY_TYPE    = re.compile(r'\b(ips|oled|tn|va|retina|nano\s*ips|amoled)\b', re.I)
RE_DISPLAY_COVER   = re.compile(r'(антиблік|антибл|матов|глян|antiglare|gloss|matte)', re.I)
RE_BRIGHTNESS      = re.compile(r'(?:яскравість|яскравость|brightness)[,:]?\s*(?:ніт|нит|nit)?\s*(\d+)|(\d+)\s*(?:ніт|нит|nit)', re.I)

RE_REFRESH_RATE    = re.compile(r'(?:частота\s*оновлення|refresh\s*rate)|(?:\d+\s*(?:гц|hz).*?(?:екран|дисплей|панель|wuxga|fhd|qhd|uhd|4k|oled|ips|tn))', re.I)
RE_CPU_FREQ        = re.compile(r'(\d+(?:[\.,]\d+)?)\s*ггц', re.I)
RE_CPU_FREQ_GHZ    = re.compile(r'(\d+(?:[\.,]\d+)?)\s*ghz', re.I)
RE_CPU_CORES       = re.compile(r'(\d+)\s*(?:ядер|ядра|cores?)', re.I)
RE_CPU_THREADS     = re.compile(r'(\d+)\s*(?:потоків|потоки|threads?)', re.I)
RE_L3_CACHE        = re.compile(r'\b(?:l3|кеш\s*l3|l3\s*кеш|cache\s*l3)\b', re.I)
RE_DDR             = re.compile(r'\b(?:LP)?DDR\d(?:X\b|\b)', re.I)
RE_RAM_FREQ        = re.compile(r'\b(\d{3,4})\s*(?:мгц|mhz)\b', re.I)
RE_RAM_SIZE        = re.compile(r'\b(\d+)\s*(?:гб|gb)\b', re.I)
RE_RAM_SLOT        = re.compile(r'(\d{1,2})\s*слот|[xх]\s*([12])(?=\s|[^0-9]|$)', re.I)
RE_STORAGE         = re.compile(r'\b(?:ssd|hdd|nvme|емкість|накоп)\b', re.I)
RE_GPU             = re.compile(r'\b(?:nvidia|geforce|rtx|gtx|radeon|intel graphics|uhd graphics|arc|iris|gpu|вбудован|дискретн|відеокарт|video)\b', re.I)
RE_OS              = re.compile(r'\b(?:windows|linux|без ос|ubuntu|dos|freeDOS)\b', re.I)
RE_BATTERY         = re.compile(r'(?:ємність|енергетична\s*ємність|battery\s*capacity)[,:]?\s*(?:(\d+)\s*)?(?:вт\*год|вт·год|wh|watt.?hour)?\s*(\d+)?', re.I)
RE_BATTERY_WHR     = re.compile(r'(\d+(?:\.\d+)?)\s*(?:Вт[\s·*\-]?год|wh\b)', re.I)
RE_CAMERA          = re.compile(r'\b(?:камера|web-камера|веб-камера|hd|4k|ir\s*camera|webcam)|\d+\s*(?:мп|mp)\b', re.I)
RE_USB_A           = re.compile(r'\b(?:usb(?:\s*type-?a|\s*3\.?\d?|\s*2\.?\d?|\s*a)|usb-a|type-?a)\b', re.I)
RE_USB_C           = re.compile(r'\b(?:usb(?:\s*type-?c|\s*c)|usb-c|type-?c|thunderbolt)\b', re.I)
RE_HDMI            = re.compile(r'\bhdmi\b', re.I)
RE_DISPLAYPORT     = re.compile(r'\b(?:displayport|dp)\b', re.I)
RE_PORTS           = re.compile(r"\b(?:порт(?:и)?|роз['']єм(?:и)?|розєм(?:и)?|interface)\b", re.I)
RE_WATERPROOF      = re.compile(r'\b(?:вологозахист|водозахист|waterproof|splash\s*proof)\b', re.I)
RE_UKRAINIAN       = re.compile(r'\b(?:україн|укр)\b', re.I)
RE_TOUCHPAD        = re.compile(r'\b(?:тачпад|маніпулятор|трекпоінт|touchpad|trackpad|numberpad)\b', re.I)
RE_3G4G            = re.compile(r'\b(?:3g|4g|lte|nano\s*sim)\b', re.I)
RE_BLUETOOTH       = re.compile(r'\bbt\b|bluetooth', re.I)
RE_WIFI            = re.compile(r'\b(?:wi-fi|wifi|wireless|wlan)\b', re.I)
RE_LAN             = re.compile(r'\b(?:lan|rj-?45|ethernet)\b', re.I)
RE_WARRANTY        = re.compile(r'гарант|warranty', re.I)

RESOLUTION_MAP = {
    "fhd":   "1920x1080",
    "wuxga": "1920x1200",
    "qhd":   "2560x1440",
    "wqxga": "2560x1600",
    "uhd":   "3840x2160",
    "4k":    "3840x2160",
}

# Values that mean "feature is explicitly absent" in spec data
_ABSENT = frozenset({
    "немає", "відсутній", "відсутня", "відсутнє", "нема",
    "no", "none", "not available", "не підтримується",
})


# ── Helper functions (same logic as Lenovo parser) ────────────────────────────

def clean_model_name(name):
    """Прибирає маркетингові суфікси з назви моделі."""
    return _NAME_SUFFIXES.sub("", name).strip()


def detect_series(name):
    nl = name.lower()
    for kw in SERIES_MAP:
        if kw.lower() in nl:
            return kw
    # За 2-символьним префіксом артикула (якщо назва — артикул)
    prefix = name[:2].upper()
    return ASUS_PN_PREFIX_SERIES.get(prefix, "-")


def extract_pn(name):
    name = name.strip()
    # ASUS part numbers like B1500CEAE-EJ3671W or 90NB0G71-M003K0
    m = re.search(r'\b([A-Z0-9]{2,4}[0-9]{3,4}[A-Z0-9_-]{4,})\b', name)
    if m:
        return m.group(1)
    candidates = RE_PART_NUMBER.findall(name)
    if candidates:
        return max(candidates, key=len)
    candidates = RE_PART_NUMBER_ALT.findall(name)
    return max(candidates, key=len) if candidates else "-"


def extract_cpu_details(cpu_model_str):
    if not cpu_model_str or cpu_model_str == "-":
        return {}
    details = {}
    s = cpu_model_str.lower()

    m = re.search(r'\((?:нов|кеш\s*)?(\d+)\s*(?:мб|mb)', s)
    if m:
        details["cache"] = m.group(1) + " МБ"

    freq_matches = re.findall(r'(\d+(?:[.,]\d+)?)\s*(?:ггц|ghz)', s)
    if freq_matches:
        freqs_numeric = []
        for f in freq_matches:
            try:
                freqs_numeric.append((float(f.replace(',', '.')), f.replace(',', '.')))
            except Exception:
                pass
        if len(freqs_numeric) >= 2:
            freqs_numeric.sort()
            details["base_freq"] = f"{freqs_numeric[0][1]} ГГц"
            details["max_freq"]  = f"{freqs_numeric[-1][1]} ГГц"
        elif len(freqs_numeric) == 1:
            if re.search(r'\b(?:до|up to|turbo|boost|max)\b', s):
                details["max_freq"] = freqs_numeric[0][1] + " ГГц"
            else:
                details["base_freq"] = freqs_numeric[0][1] + " ГГц"

    cores = re.search(r'(\d+)(?:\s*-)?core|(\d+)\s*ядер', s)
    if cores:
        details["cores"] = cores.group(1) or cores.group(2)

    threads = re.search(r'(\d+)(?:\s*-)?thread|(\d+)\s*потоків', s)
    if threads:
        details["threads"] = threads.group(1) or threads.group(2)

    return details


def parse_specs(items):
    keys = [
        "CPU_Brand", "CPU_Model", "CPU_Cores", "CPU_Threads",
        "CPU_Base_Freq_GHz", "CPU_Max_Freq_GHz", "CPU_Cache_L3_MB",
        "RAM_Type", "RAM_Freq", "RAM_Size_GB", "RAM_Slots",
        "Nakopychuvach_SSD", "Display_Diagonal", "Display_Max_Resolution",
        "Display_Matrix_Type", "Display_Cover", "Display_Brightness_nits",
        "Display_Refresh_Rate",
        "Video_Brand", "GPU_Type", "GPU_Model", "GPU_Memory_MB",
        "OS", "Battery_Capacity_Wh", "Camera_MP", "USB_TypeA", "USB_TypeC",
        "HDMI", "Keyboard_Waterproof",
        "Keyboard_Ukrainian", "Keyboard_Pointing_Device", "Network_3G4G",
        "Bluetooth", "WiFi", "LAN_Mbps", "Certificates", "Warranty",
        "Merezha_WiFi",
    ]
    r = {k: "-" for k in keys}

    def set_if_empty(key, value):
        if r[key] == "-" and value:
            r[key] = value.rstrip(";").strip() if isinstance(value, str) else str(value)

    def set_from_detail(key, value):
        if value and (r[key] in ("-", "Так", "Є")):
            r[key] = value.rstrip(";").strip() if isinstance(value, str) else str(value)

    full_text = " ".join([str(item).lower() for item in items if item])

    # First pass: DDR type
    for s in items:
        if not s:
            continue
        if RE_DDR.search(s):
            m = RE_DDR.search(s)
            if m:
                r["RAM_Type"] = m.group(0)
                break

    for s in items:
        if not s:
            continue
        sl = s.lower()

        # ── Display ──────────────────────────────────────────────────────────
        if RE_DIAG.search(sl):
            m = RE_DIAG.search(sl)
            if m and r["Display_Diagonal"] == "-":
                set_if_empty("Display_Diagonal", m.group(0))

        if RE_RESOLUTION.search(sl):
            m = RE_RESOLUTION.search(sl)
            if m:
                pixel_m = re.search(r'(\d{3,4})[xх](\d{3,4})', sl)
                name_m  = re.search(r'\b(fhd|wuxga|qhd|wqxga|uhd|4k)\b', sl, re.I)
                if pixel_m and name_m:
                    set_if_empty("Display_Max_Resolution",
                                 f"{name_m.group(0).upper()} ({pixel_m.group(0).upper()})")
                elif pixel_m:
                    set_if_empty("Display_Max_Resolution", pixel_m.group(0).upper())
                elif name_m:
                    name_key = name_m.group(0).lower()
                    pixels = RESOLUTION_MAP.get(name_key, "")
                    if pixels:
                        set_if_empty("Display_Max_Resolution",
                                     f"{name_m.group(0).upper()} ({pixels})")
                    else:
                        set_if_empty("Display_Max_Resolution", name_m.group(0).upper())
                else:
                    set_if_empty("Display_Max_Resolution", m.group(0).upper())

        if RE_DISPLAY_TYPE.search(sl):
            m = RE_DISPLAY_TYPE.search(sl)
            if m:
                set_if_empty("Display_Matrix_Type", m.group(0).upper())

        if RE_DISPLAY_COVER.search(sl):
            cover_m = RE_DISPLAY_COVER.search(sl)
            if cover_m:
                ct = cover_m.group(0).lower()
                if "глян" in ct or "gloss" in ct:
                    set_if_empty("Display_Cover", "Глянцеве покриття")
                elif "матов" in ct or "matte" in ct:
                    set_if_empty("Display_Cover", "Матове покриття")
                elif "антиблі" in ct or "antiglare" in ct:
                    set_if_empty("Display_Cover", "Антиблікове покриття")
                else:
                    set_if_empty("Display_Cover", ct)

        if RE_BRIGHTNESS.search(sl):
            m = RE_BRIGHTNESS.search(sl)
            if m:
                val = m.group(1) or m.group(2)
                if val:
                    set_if_empty("Display_Brightness_nits", f"{val} nits")

        if RE_REFRESH_RATE.search(sl):
            freq_m = re.search(r'(\d+)\s*(?:hz|гц)\b', sl, re.I)
            if freq_m:
                set_if_empty("Display_Refresh_Rate", freq_m.group(0))
            else:
                freq_m = re.search(r'(?:частота\s*оновлення|refresh\s*rate)\s*[:\s]+(\d+)', sl, re.I)
                if freq_m:
                    set_if_empty("Display_Refresh_Rate", freq_m.group(1) + " Hz")
        # Hz on a display-context line where keyword appears BEFORE Hz (e.g. "FHD 144Hz", "WQXGA 240Hz")
        if r["Display_Refresh_Rate"] == "-":
            if RE_RESOLUTION.search(sl) or RE_DISPLAY_TYPE.search(sl) or RE_DIAG.search(sl):
                hz_m = re.search(r'\b(\d{2,3})\s*(?:hz|гц)\b', sl, re.I)
                if hz_m:
                    val = int(hz_m.group(1))
                    if 30 <= val <= 480:
                        set_if_empty("Display_Refresh_Rate", hz_m.group(0))

        # ── CPU ──────────────────────────────────────────────────────────────
        if r["CPU_Brand"] == "-":
            if "intel" in sl:
                set_if_empty("CPU_Brand", "Intel")
            elif "amd" in sl or "ryzen" in sl:
                set_if_empty("CPU_Brand", "AMD")
            elif "snapdragon" in sl:
                set_if_empty("CPU_Brand", "Snapdragon")
            elif "mediatek" in sl:
                set_if_empty("CPU_Brand", "MediaTek")

        cpu_related = any(k in sl for k in [
            "intel", "amd", "ryzen", "snapdragon", "celeron", "pentium",
            "core", "xeon", "i3", "i5", "i7", "i9", "процес", "cpu", "mediatek"
        ])
        if cpu_related and r["CPU_Model"] == "-":
            clean = re.sub(r'\s*\([^)]*(?:мб|ггц|кеш|ghz|mb|cache)[^)]*\)', '', s, flags=re.I).strip()
            clean = re.sub(r'\s*[®™]\s*', ' ', clean).strip()
            cl = clean.lower()
            if "intel" in cl:
                # Patterns: "Core i7-13700H" / "Core Ultra 9 Processor 285H" / "Core Ultra X9"
                m = re.search(
                    r'Core(?:\s+Ultra)?\s+'
                    r'(?:i[3-9][\w-]+|'                              # i3/i5/i7/i9
                    r'\d+\w*(?:\s+\w+){0,2}(?:\s+\d{3,4}[A-Z]{1,3})?|'  # "9 Processor 285H"
                    r'[A-Z]\w*(?:\s+\w*){0,2})',                     # "X9 Series 3"
                    clean, re.I)
                if not m:
                    m = re.search(r'(?:Celeron|Pentium|Xeon)\s+[\w-]+', clean, re.I)
                set_if_empty("CPU_Model", m.group(0).strip() if m else clean)
            elif "amd" in cl or "ryzen" in cl:
                m = re.search(r'Ryzen\s+\d+\s+\d{4,5}\w*', clean, re.I)
                if not m:
                    m = re.search(r'(?:Ryzen|Athlon)\s+\w+(?:\s+\w+)?', clean, re.I)
                set_if_empty("CPU_Model", m.group(0).strip() if m else clean)
            elif "snapdragon" in cl:
                m = re.search(r'Snapdragon\s+\w+(?:\s+[\w-]+)+', clean, re.I)
                if m:
                    set_if_empty("CPU_Model", m.group(0).strip())
            else:
                set_if_empty("CPU_Model", clean)

        if RE_CPU_CORES.search(sl):
            m = RE_CPU_CORES.search(sl)
            if m:
                set_if_empty("CPU_Cores", m.group(1))

        if RE_CPU_THREADS.search(sl):
            m = RE_CPU_THREADS.search(sl)
            if m:
                set_if_empty("CPU_Threads", m.group(1))

        if cpu_related:
            base_freq_m = re.search(r'(?:номін|базов)\w*\s*частота[\s,ггц]*(\d+(?:[.,]\d+)?)', sl, re.I)
            max_freq_m  = re.search(r'максимальна\s*частота[\s,ггц]*(\d+(?:[.,]\d+)?)', sl, re.I)
            if base_freq_m:
                set_if_empty("CPU_Base_Freq_GHz", base_freq_m.group(1) + " ГГц")
            if max_freq_m:
                set_if_empty("CPU_Max_Freq_GHz", max_freq_m.group(1) + " ГГц")

            freq_matches = RE_CPU_FREQ.findall(sl)
            if not freq_matches:
                freq_matches = RE_CPU_FREQ_GHZ.findall(sl)

            if len(freq_matches) >= 2:
                freqs = [float(f.replace(',', '.')) for f in freq_matches]
                low, high = min(freqs), max(freqs)
                set_if_empty("CPU_Base_Freq_GHz", f"{low} ГГц")
                set_if_empty("CPU_Max_Freq_GHz",  f"{high} ГГц")
            elif len(freq_matches) == 1:
                if "до" in sl or "turbo" in sl or "boost" in sl or "max" in sl or "up to" in sl:
                    set_if_empty("CPU_Max_Freq_GHz", freq_matches[0] + " ГГц")
                else:
                    set_if_empty("CPU_Base_Freq_GHz", freq_matches[0] + " ГГц")

        if RE_L3_CACHE.search(sl):
            m = re.search(r'l3[^0-9\n]{0,20}?(\d+)\s*(?:мб|mb)?', sl, re.I)
            if m:
                set_if_empty("CPU_Cache_L3_MB", m.group(1) + " МБ")
        # ASUS format: "(кеш 24 МБ)" or "кеш 24 МБ" inside CPU spec
        m = re.search(r'\bкеш\s+(\d+)\s*(?:мб|mb)\b', sl, re.I)
        if m:
            set_if_empty("CPU_Cache_L3_MB", m.group(1) + " МБ")

        # ── RAM ──────────────────────────────────────────────────────────────
        if RE_RAM_FREQ.search(sl):
            m = RE_RAM_FREQ.search(sl)
            if m:
                set_if_empty("RAM_Freq", m.group(1) + " МГц")
        # DDR speed without unit: "DDR5-4800", "DDR5 4800", "LPDDR5X-9600" — ≥2133 = valid
        if r["RAM_Freq"] == "-" and RE_DDR.search(sl):
            m = re.search(r'(?:lp)?ddr\d[xх]?[\s\-–](\d{4,5})(?:\s*(?:мгц|mhz))?', sl, re.I)
            if m and int(m.group(1)) >= 2133:
                set_if_empty("RAM_Freq", m.group(1) + " МГц")
        # Fallback: MHz value on a RAM-context line ("Швидкість пам'яті 4800 МГц")
        if r["RAM_Freq"] == "-":
            if any(k in sl for k in ["пам", "швидкіст", "частот", "memory speed", "ram speed"]):
                m = re.search(r'\b(\d{4,5})\s*(?:мгц|mhz)\b', sl, re.I)
                if m and int(m.group(1)) >= 2133:
                    set_if_empty("RAM_Freq", m.group(1) + " МГц")

        if (r["RAM_Size_GB"] == "-"
                and "ssd" not in sl and "hdd" not in sl and "nvme" not in sl
                and "дисплей" not in sl and "display" not in sl):
            m = RE_RAM_SIZE.search(sl)
            if m:
                set_if_empty("RAM_Size_GB", m.group(0).upper())

        if "слот" in sl and any(k in sl for k in ["озу", "ddr", "пам", "гб", "ram"]):
            m = RE_RAM_SLOT.search(sl)
            if m:
                slot_num = m.group(1) or m.group(2)
                if slot_num:
                    set_if_empty("RAM_Slots", slot_num + " слотів")
        # ASUS "Слоти розширення: 1x DDR5 SO-DIMM ..." — digit before x
        if 'so-dimm' in sl:
            m = re.search(r'(\d+)[xх]\s*(?:lp)?ddr\d\s*so-dimm', sl, re.I)
            if m:
                set_if_empty("RAM_Slots", m.group(1))

        # ── Storage ──────────────────────────────────────────────────────────
        if RE_STORAGE.search(sl):
            # Pattern 1: "SSD до 2 ТБ ..." or "накопичувач 2 ТБ ..." (compare_row format)
            _sm = re.search(
                r'(?:ssd\s+до|накопичувач)\s+(\d+(?:[.,]\d+)?\s*(?:тб|гб|tb|gb)[^‖\n]{0,60})',
                s, re.I)
            if _sm:
                _sv = re.sub(r'[™®©]', '', _sm.group(1)).strip().rstrip(',').strip()
                set_if_empty("Nakopychuvach_SSD", re.sub(r'\s+', ' ', _sv))
            else:
                # Pattern 2: "2TB M.2 5.0 NVMe PCIe 4.0 SSD" inside a long combined bullet
                _sm = re.search(
                    r'(\d+(?:[.,]\d+)?)\s*(?:тб|гб|tb|gb)'
                    r'(?:\s+m\.?2(?:\s+[\d.]+)?)?'
                    r'(?:\s+nvme(?:™)?)?'
                    r'(?:\s+pcie(?:®)?(?:\s*[\d.]+)?(?:\s*x\d+)?)?'
                    r'(?:\s+(?:performance|high[\s-]speed))?'
                    r'(?:\s+(?:ssd|hdd))?',
                    s, re.I)
                if _sm:
                    _sv = re.sub(r'[™®©]', '', _sm.group(0)).strip()
                    set_if_empty("Nakopychuvach_SSD", re.sub(r'\s+', ' ', _sv))
                elif len(s) < 100:
                    set_if_empty("Nakopychuvach_SSD", s.strip())

        # ── GPU ──────────────────────────────────────────────────────────────
        if RE_GPU.search(sl):
            sl_lower = sl.lower()

            if r["GPU_Memory_MB"] == "-":
                # Explicit VRAM keyword, OR discrete GPU (RTX/GTX/Radeon RX) — VRAM size 2-24 GB
                vram_kw = re.search(r'gddr|vram|відеопам|gpu\s*memory|обсяг\s*gpu', sl, re.I)
                is_discrete = any(k in sl_lower for k in ["rtx", "gtx", "radeon rx"])
                if vram_kw or is_discrete:
                    m = re.search(r'(\d+)\s*(?:гб|gb)', sl, re.I)
                    if m:
                        val = int(m.group(1))
                        if 2 <= val <= 24:
                            set_if_empty("GPU_Memory_MB", str(val) + " GB")

            if r["Video_Brand"] == "-":
                if any(k in sl_lower for k in ["nvidia", "geforce", "rtx", "gtx"]):
                    set_if_empty("Video_Brand", "NVIDIA")
                elif any(k in sl_lower for k in ["amd", "radeon"]):
                    set_if_empty("Video_Brand", "AMD")
                elif any(k in sl_lower for k in ["intel", "iris", "arc"]):
                    set_if_empty("Video_Brand", "Intel")
                elif any(k in sl_lower for k in ["qualcomm", "adreno"]):
                    set_if_empty("Video_Brand", "Qualcomm")

            if r["GPU_Model"] == "-":
                m = re.search(r'\b(?:rtx|gtx|radeon\s+rx)\s*(\d{3,4})\b', sl, re.I)
                if m:
                    prefix = re.search(r'(rtx|gtx|radeon\s+rx)', sl, re.I)
                    set_if_empty("GPU_Model", f"{prefix.group(0).upper()} {m.group(1)}")
                else:
                    m = re.search(
                        r'\b(?:radeon\s+\d{3}m|arc\s+\w+|iris\s+\w+|intel\s+graphics|'
                        r'(?:intel\s+)?(?:uhd|hd|iris)\s+graphics|(?:qualcomm\s+)?adreno\s+gpu)\b',
                        sl, re.I)
                    if m:
                        model_text = m.group(0).strip()
                        parts = model_text.split()
                        model_text = ' '.join(
                            [p.upper() if p.upper() in ('UHD', 'HD', 'XE') else p.title()
                             for p in parts])
                        set_if_empty("GPU_Model", model_text)
                    else:
                        brand_m = re.search(
                            r'(?:nvidia|geforce|amd|radeon|intel|arc|qualcomm|adreno)', sl, re.I)
                        if brand_m:
                            gpu_text = s[brand_m.start():].strip()
                            gpu_text = re.sub(r',.*', '', gpu_text).strip()
                            gpu_text = re.sub(r'\d+\s*(?:гб|gb|мб|mb)', '', gpu_text, flags=re.I).strip()
                            gpu_text = re.sub(r'\s*\([^)]*\)', '', gpu_text).strip()
                            gpu_text = re.sub(r'[™®©]', '', gpu_text).strip()  # remove trademark symbols
                            gpu_text = re.sub(r'\s+', ' ', gpu_text).strip()
                            if 3 < len(gpu_text) < 60:
                                set_if_empty("GPU_Model", gpu_text.title())

            if r["GPU_Type"] == "-":
                sl_norm = re.sub(r'[™®©]', '', sl_lower)  # strip trademark symbols for pattern matching
                if r["Video_Brand"] == "NVIDIA" or any(k in sl_norm for k in ["rtx", "gtx", "radeon rx"]):
                    set_if_empty("GPU_Type", "дискретна")
                elif r["Video_Brand"] in ("Intel", "Qualcomm") or any(
                        k in sl_norm for k in ["iris", "arc", "uhd graphics", "adreno"]):
                    set_if_empty("GPU_Type", "інтегрована")
                elif r["Video_Brand"] == "AMD":
                    if any(k in sl_norm for k in ["radeon 780m", "radeon 680m", "radeon graphics"]) \
                            or re.search(r'radeon\s+\d{3,4}m\b', sl_norm):
                        set_if_empty("GPU_Type", "інтегрована")
                    elif re.search(r'radeon\s+rx', sl_norm):
                        set_if_empty("GPU_Type", "дискретна")
                    else:
                        set_if_empty("GPU_Type", "дискретна")

        # ── OS ───────────────────────────────────────────────────────────────
        if RE_OS.search(sl):
            win_m = re.search(r'Windows\s+\d+\s+\w+', s, re.I)
            if win_m:
                set_if_empty("OS", win_m.group(0))
            else:
                m = RE_OS.search(sl)
                val = m.group(0).strip()
                set_if_empty("OS", val.upper() if val.lower() in ('без ос', 'dos', 'freedos') else val.title())

        # ── Battery ──────────────────────────────────────────────────────────
        whr_m = RE_BATTERY_WHR.search(s)
        if whr_m:
            set_if_empty("Battery_Capacity_Wh", whr_m.group(1) + " Wh")
        elif RE_BATTERY.search(sl):
            m = RE_BATTERY.search(sl)
            if m:
                val = m.group(1) or m.group(2)
                if val:
                    set_if_empty("Battery_Capacity_Wh", val + " Wh")
        # Fallback: numeric Wh value on a battery-context line
        if r["Battery_Capacity_Wh"] == "-":
            if any(k in sl for k in ["батар", "battery", "акумул", "ємність"]):
                m = re.search(r'(\d+(?:\.\d+)?)\s*(?:вт|wh|watt)', sl, re.I)
                if m:
                    try:
                        bval = float(m.group(1))
                        if 20 <= bval <= 200:
                            set_if_empty("Battery_Capacity_Wh", m.group(1) + " Wh")
                    except ValueError:
                        pass

        # ── Camera ───────────────────────────────────────────────────────────
        if RE_CAMERA.search(sl):
            cam_m = re.search(r'(\d+(?:\.\d+)?)\s*(?:мп|mp)', sl, re.I)
            if cam_m:
                set_if_empty("Camera_MP", cam_m.group(0).upper())
            elif re.search(r'\b1080p?\b', sl) and ("camera" in sl or "камера" in sl or "webcam" in sl):
                set_if_empty("Camera_MP", "FHD (1080p)")
            elif re.search(r'\b720p?\b', sl) and ("camera" in sl or "камера" in sl or "webcam" in sl):
                set_if_empty("Camera_MP", "HD (720p)")
            elif "fhd" in sl and ("camera" in sl or "камера" in sl or "webcam" in sl):
                set_if_empty("Camera_MP", "FHD")
            elif "4k" in sl and ("camera" in sl or "камера" in sl or "webcam" in sl):
                set_if_empty("Camera_MP", "4K")
            elif "hd" in sl and ("camera" in sl or "камера" in sl or "webcam" in sl):
                set_if_empty("Camera_MP", "HD")

        # ── Ports / Keyboard / Network — split on ‖ separator ────────────────
        if "‖" in s:
            usb_a_parts, usb_c_parts = [], []
            for entry in s.split("‖"):
                e  = entry.strip()
                el = e.lower()

                if re.match(r'процесор\s', el) or re.match(r'cpu\s', el):
                    raw = re.sub(r'^(?:процесор|cpu)\s*', '', e, flags=re.I).strip()
                    raw = re.sub(r'\s*\([^)]*(?:мб|ггц|кеш|ghz|mb|cache)[^)]*\)', '', raw, flags=re.I).strip()
                    raw = re.sub(r'\s*[®™]\s*', ' ', raw).strip()
                    # Strip leading navigation markers like "– до" or "до"
                    raw = re.sub(r'^[–\-]\s*до\s+', '', raw).strip()
                    raw = re.sub(r'^до\s+', '', raw, flags=re.I).strip()
                    if raw and r["CPU_Model"] == "-":
                        ms = re.search(r'Snapdragon\s+\w+(?:\s+[\w-]+)+', raw, re.I)
                        mi = re.search(
                            r'Core(?:\s+Ultra)?\s+'
                            r'(?:i[3-9][\w-]+|'
                            r'\d+\w*(?:\s+\w+){0,2}(?:\s+\d{3,4}[A-Z]{1,3})?|'
                            r'[A-Za-z]\w*(?:\s+\w+){0,2})',
                            raw, re.I)
                        ma = re.search(r'Ryzen\s+(?:AI\s+)?(?:\w+\s+)?\w+', raw, re.I)
                        if not ma:
                            ma = re.search(r'(?:Ryzen|Athlon)\s+\w+(?:\s+\w+)?', raw, re.I)
                        m  = ms or mi or ma
                        set_if_empty("CPU_Model", m.group(0).strip() if m else raw)

                elif re.match(r'(?:кеш\s*l3|l3\s*кеш)', el):
                    val = re.sub(r'^(?:кеш\s*l3|l3\s*кеш)\s*', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("CPU_Cache_L3_MB", val)

                elif re.match(r"тип\s*пам", el) or re.match(r"memory\s*type", el):
                    m = re.search(r'\b(?:LP)?DDR\d(?:X\b|\b)', e, re.I)
                    if m:
                        set_if_empty("RAM_Type", m.group(0).upper())

                elif re.match(r'usb\s*type-?a\s', el) or re.match(r'usb\s*3', el):
                    val = re.sub(r'^usb[\s\w-]*\s*', '', e, flags=re.I).strip()
                    if val:
                        usb_a_parts.append(val)

                elif re.match(r'usb[\s-]*type-?c\s', el) or re.match(r'thunderbolt', el):
                    val = re.sub(r'^(?:usb[\s-]*type-?c|thunderbolt)\s*', '', e, flags=re.I).strip()
                    if val:
                        usb_c_parts.append(val)

                elif re.match(r'hdmi[,\s]', el):
                    val = re.sub(r'^hdmi[,\s]*(?:шт\.?\s+)?', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("HDMI", val)

                elif re.match(r'клавіатура\s*\(вологозахист\)', el) or re.match(r'splash\s*proof', el):
                    val = re.sub(r'^(?:клавіатура\s*\(вологозахист\)|splash\s*proof)\s*', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("Keyboard_Waterproof", val)

                elif re.match(r'українська\s*мова\s', el):
                    val = re.sub(r'^українська\s*мова\s*', '', e, flags=re.I).strip()
                    if val:
                        set_if_empty("Keyboard_Ukrainian", val)

                elif re.match(r'маніпулятори\s', el) or re.match(r'touchpad', el):
                    val = re.sub(r'^(?:маніпулятори|touchpad)\s*', '', e, flags=re.I).strip()
                    if val:
                        if val.lower() in _ABSENT:
                            set_if_empty("Keyboard_Pointing_Device", "Нема")
                        else:
                            set_if_empty("Keyboard_Pointing_Device", val)

                elif re.match(r'3g/4g\s', el) or re.match(r'lte', el):
                    val = re.sub(r'^(?:3g/4g|lte)\s*', '', e, flags=re.I).strip()
                    if val:
                        if val.lower() in _ABSENT:
                            set_from_detail("Network_3G4G", "Нема")
                        else:
                            set_from_detail("Network_3G4G", val)

                elif re.match(r'bluetooth\s', el):
                    val = re.sub(r'^bluetooth\s*', '', e, flags=re.I).strip()
                    val = re.sub(r'[®™]', '', val).strip()
                    if val:
                        if val.lower() in _ABSENT:
                            set_from_detail("Bluetooth", "Нема")
                        else:
                            bt_val = ("Bluetooth " + val) if re.match(r'[\d.]', val) else val
                            set_from_detail("Bluetooth", bt_val)

                elif re.match(r'wi-fi\s', el) or re.match(r'wlan\s', el) or re.match(r'wireless\s', el):
                    val = re.sub(r'^(?:wi-fi|wlan|wireless)\s*', '', e, flags=re.I).strip()
                    val = re.sub(r'[®™]', '', val).strip()
                    if val:
                        if val.lower() in _ABSENT:
                            set_from_detail("WiFi", "Нема")
                            set_from_detail("Merezha_WiFi", "Нема")
                        else:
                            wifi_val = ("Wi-Fi " + val) if not re.match(r'wi-?fi', val, re.I) else val
                            set_from_detail("WiFi", wifi_val)
                            set_from_detail("Merezha_WiFi", wifi_val)

                elif re.match(r'lan\s*rj-?45', el) or re.match(r'ethernet', el):
                    val = re.sub(r'^(?:lan\s*rj-?45|ethernet)[,\s]*(?:мбіт/с\s*)?', '', e, flags=re.I).strip()
                    if val:
                        if val.lower() in _ABSENT:
                            set_from_detail("LAN_Mbps", "Нема")
                        else:
                            set_from_detail("LAN_Mbps", val)

                elif (re.match(r'(?:екологічні\s*)?сертифікат', el)
                      or re.match(r'certificate', el)
                      or re.match(r'військові\s*стандарт', el)):
                    val = re.sub(
                        r'^(?:екологічні\s*)?(?:сертифікат\w*|certificate\w*|військові\s*стандарт\w*)\s*',
                        '', e, flags=re.I).strip()
                    if val:
                        # Combine with existing value if already set
                        existing = r.get("Certificates", "-")
                        if existing not in ("-", "", val) and val not in existing:
                            set_from_detail("Certificates", existing + " | " + val)
                        else:
                            set_from_detail("Certificates", val)

                elif re.match(r'(?:термін\s*базово|warranty)', el):
                    m = re.search(r'(\d+\s*(?:рік|роки|років|місяц\w*|year|month))', e, re.I)
                    if m:
                        set_from_detail("Warranty", m.group(1).strip())

            if usb_a_parts:
                set_if_empty("USB_TypeA", " / ".join(usb_a_parts))
            if usb_c_parts:
                set_if_empty("USB_TypeC", " / ".join(usb_c_parts))
        else:
            # Fallback: no ‖ separator
            has_ports_kw = RE_PORTS.search(sl)
            has_usb_a    = RE_USB_A.search(sl)
            has_usb_c    = RE_USB_C.search(sl)
            has_hdmi     = RE_HDMI.search(sl)
            multi_port   = sum(map(bool, [has_usb_a, has_usb_c, has_hdmi])) > 1

            if has_ports_kw or multi_port:
                m = re.search(r'\d+(?:x|\s+порт\w*)?\s+USB[\s\d.]+(?:Gen\s*\d+\s+)?Type-?A', s, re.I)
                if m:
                    set_if_empty("USB_TypeA", m.group(0).strip())

                mc = re.search(r'\d+(?:x|\s+порт\w*)?\s+USB[\s\d.]+(?:Gen\s*\d+\s+)?Type-?C(?:\s+з\s+підтримкою[^,\d\n]*)?', s, re.I)
                tb = re.search(r'Thunderbolt\s+\d+[^,\d\n]*', s, re.I)
                usb_c_val = mc.group(0).strip() if mc else ""
                tb_val    = tb.group(0).strip() if tb else ""
                combined_c = " / ".join(filter(None, [usb_c_val, tb_val]))
                if combined_c:
                    set_if_empty("USB_TypeC", combined_c)

                m = re.search(r'\d+(?:x|\s+порт\w*)?\s+HDMI\s+[\d.]+(?:\s*,\s*TMDS)?', s, re.I)
                if m:
                    set_if_empty("HDMI", m.group(0).strip())

            else:
                # Single-type line
                if has_usb_a:
                    set_if_empty("USB_TypeA", s.strip())
                if has_usb_c:
                    set_if_empty("USB_TypeC", s.strip())
                if has_hdmi:
                    set_if_empty("HDMI", s.strip())
            _line_absent = any(k in sl for k in ("немає", "відсутній", "відсутня", "нема"))
            if RE_WATERPROOF.search(sl):
                set_if_empty("Keyboard_Waterproof", "Нема" if _line_absent else "Так")
            if RE_UKRAINIAN.search(sl):
                set_if_empty("Keyboard_Ukrainian", "Нема" if _line_absent else "Так")
            if RE_TOUCHPAD.search(sl):
                set_if_empty("Keyboard_Pointing_Device", "Нема" if _line_absent else "Так")
            if RE_3G4G.search(sl):
                set_if_empty("Network_3G4G", "Нема" if _line_absent else "Так")
            if RE_BLUETOOTH.search(sl):
                if _line_absent:
                    set_if_empty("Bluetooth", "Нема")
                else:
                    bt_m = re.search(r'Bluetooth[\s®™]*(\d+(?:\.\d+)?(?:\s*\+\s*LE)?)', s, re.I)
                    set_if_empty("Bluetooth", ("Bluetooth " + bt_m.group(1).strip()) if bt_m else "Так")
            if RE_WIFI.search(sl):
                if _line_absent:
                    set_if_empty("WiFi", "Нема")
                    set_if_empty("Merezha_WiFi", "Нема")
                else:
                    wifi_m = re.search(r'Wi-?Fi\s*(?:\d+[eE]?\s*(?:\(802\.\d+[a-z]*\))?)', s, re.I)
                    wifi_val = re.sub(r'\s+', ' ', re.sub(r'[®™]', '', wifi_m.group(0))).strip() if wifi_m else "Так"
                    set_if_empty("WiFi", wifi_val)
                    set_if_empty("Merezha_WiFi", wifi_val)
            if RE_LAN.search(sl):
                if _line_absent:
                    set_if_empty("LAN_Mbps", "Нема")
                elif "gigabit" in sl or "гігабіт" in sl:
                    set_if_empty("LAN_Mbps", "1000 Мбіт/с")
                else:
                    m = re.search(r'(\d{3,4})\s*(?:мбіт|mbps)', sl, re.I)
                    if m:
                        set_if_empty("LAN_Mbps", m.group(0))
                    else:
                        set_if_empty("LAN_Mbps", "Так")

        # ── Warranty ─────────────────────────────────────────────────────────
        if RE_WARRANTY.search(sl) or "термін" in sl:
            m = re.search(r'(\d+)\s*(?:місяц\w*|рік\w*|роки|років|year\w*|month\w*)', sl, re.I)
            if m:
                set_if_empty("Warranty", m.group(0))
            elif RE_WARRANTY.search(sl):
                set_if_empty("Warranty", "Так")
        # ── Certificates (non-‖ path) ─────────────────────────────────────────
        if ("сертифікат" in sl or "certificate" in sl
                or "mil-std" in sl or "energy star" in sl or "epeat" in sl
                or "rohs" in sl):
            cert_m = re.search(
                r'(?:EPEAT|Energy\s+Star|RoHS|REACH|MIL-STD|FCC|CE\b|BSMI)[^\n]{0,120}',
                s, re.I)
            if cert_m:
                cert_val = cert_m.group(0).strip()
                existing = r.get("Certificates", "-")
                if existing not in ("-", "", cert_val) and cert_val not in existing:
                    set_if_empty("Certificates", existing + " | " + cert_val)
                else:
                    set_if_empty("Certificates", cert_val)

    # Post-process CPU details from model string
    if r["CPU_Model"] != "-":
        cpu_details = extract_cpu_details(r["CPU_Model"])
        for k, v in [
            ("CPU_Cache_L3_MB", cpu_details.get("cache")),
            ("CPU_Base_Freq_GHz", cpu_details.get("base_freq")),
            ("CPU_Max_Freq_GHz", cpu_details.get("max_freq")),
            ("CPU_Cores", cpu_details.get("cores")),
            ("CPU_Threads", cpu_details.get("threads")),
        ]:
            if v:
                set_if_empty(k, v)

    # Fill missing CPU details from lookup table (for ROG / pages without compare_rows)
    if r["CPU_Model"] != "-" and any(
            r[k] == "-" for k in ("CPU_Cores", "CPU_Threads", "CPU_Base_Freq_GHz",
                                   "CPU_Max_Freq_GHz", "CPU_Cache_L3_MB")):
        model_str = r["CPU_Model"]
        for key, (cores, threads, base, top, cache) in CPU_SPECS.items():
            if re.search(re.escape(key), model_str, re.I):
                set_if_empty("CPU_Cores",        str(cores))
                set_if_empty("CPU_Threads",      str(threads))
                set_if_empty("CPU_Base_Freq_GHz", base + " ГГц")
                set_if_empty("CPU_Max_Freq_GHz",  top  + " ГГц")
                set_if_empty("CPU_Cache_L3_MB",   cache + " МБ")
                break

    # Search cores/threads in full text if still missing
    if r["CPU_Cores"] == "-" or r["CPU_Threads"] == "-":
        cores_m   = re.search(r'кількість ядер\s+(\d+)', full_text, re.I)
        threads_m = re.search(r'кількість потоків\s+(\d+)', full_text, re.I)
        if cores_m:
            set_if_empty("CPU_Cores", cores_m.group(1))
        if threads_m:
            set_if_empty("CPU_Threads", threads_m.group(1))
        if not cores_m:
            cores_m = re.search(r'(\d+)\s*(?:ядер|ядра|cores?)', full_text, re.I)
            if cores_m:
                set_if_empty("CPU_Cores", cores_m.group(1))
        if not threads_m:
            threads_m = re.search(r'(\d+)\s*(?:потоків|потоки|threads?)', full_text, re.I)
            if threads_m:
                set_if_empty("CPU_Threads", threads_m.group(1))

    # Post-process CPU_Model: strip leftover navigation/description prefixes and suffixes
    if r["CPU_Model"] not in ("-",):
        cm = r["CPU_Model"]
        cm = re.sub(r'^[–\-]\s*до\s+', '', cm).strip()       # "– до Intel ..." -> "Intel ..."
        cm = re.sub(r'^до\s+', '', cm, flags=re.I).strip()    # "до Intel ..." -> "Intel ..."
        cm = re.sub(r'\s+з\s+(?:графікою|відеокартою).*$', '', cm, flags=re.I).strip()  # strip " з графікою..."
        cm = re.sub(r'\s+(?:Series\s+\d+)$', '', cm, flags=re.I).strip()  # strip trailing "Series 3"
        cm = re.sub(r'\s+\([^)]*\)\s*$', '', cm).strip()      # strip trailing (...) if remains
        # Strip Intel marketing word "Processor": "Core Ultra 7 Processor 255H" -> "Core Ultra 7 255H"
        cm = re.sub(r'\s+Processor\b', '', cm, flags=re.I).strip()
        # Strip display-size digit glued to Intel SKU suffix: "275HX18" -> "275HX", "386H14" -> "386H"
        cm = re.sub(r'(\b\d{3,4}[A-Z]{1,3})\d{1,2}\b', r'\1', cm)
        if cm and cm != r["CPU_Model"]:
            r["CPU_Model"] = cm
        # If model still looks like full marketing text (contains "до" at start), try re-extract
        if re.match(r'^(?:процесор|до\b)', r["CPU_Model"], re.I):
            m2 = re.search(r'(?:Core(?:\s+Ultra)?\s+\w+(?:\s+\w+){0,2}|Ryzen(?:\s+AI)?\s+\w+(?:\s+\w+)?|Snapdragon\s+\w+(?:\s+[\w-]+)+)', r["CPU_Model"], re.I)
            if m2:
                r["CPU_Model"] = m2.group(0).strip()

    # GPU_Memory: integrated GPU never has dedicated VRAM
    if r["GPU_Memory_MB"] == "-" and r["GPU_Type"] == "інтегрована":
        r["GPU_Memory_MB"] = "Нема"

    # 3G/4G: if we have network info (WiFi/BT found) but no 3G/4G row -> explicitly absent
    has_network_data = r["WiFi"] not in ("-",) or r["Bluetooth"] not in ("-",)
    if r["Network_3G4G"] == "-" and has_network_data:
        r["Network_3G4G"] = "Нема"

    # LAN: thin/consumer laptops without LAN spec -> absent
    # ExpertBook, ROG, TUF have LAN ports — exclude them
    _lan_series = any(k in full_text for k in ["expertbook", "rog", "tuf"])
    if r["LAN_Mbps"] == "-" and has_network_data and not _lan_series:
        r["LAN_Mbps"] = "Нема"

    # ASUS default assumptions
    if r["Bluetooth"] == "-" and "asus" in full_text:
        set_if_empty("Bluetooth", "Так")
    if r["WiFi"] == "-" and "asus" in full_text:
        set_if_empty("WiFi", "Так")
        set_if_empty("Merezha_WiFi", "Так")

    return r


# ── ASUS detail-page extraction ───────────────────────────────────────────────

def _build_record(series_name, model_num, price, buy_url, spec_items):
    """Build a full record dict from spec items list."""
    clean_name = clean_model_name(series_name)
    specs = parse_specs(spec_items)
    return {
        "Seria":             detect_series(clean_name),
        "Model":             clean_name,
        "Part Number":       model_num if model_num != "-" else extract_pn(clean_name),
        "CPU_Brand":         specs["CPU_Brand"],
        "CPU_Model":         specs["CPU_Model"],
        "CPU_Cores":         specs["CPU_Cores"],
        "CPU_Threads":       specs["CPU_Threads"],
        "CPU_Base_Freq_GHz": specs["CPU_Base_Freq_GHz"],
        "CPU_Max_Freq_GHz":  specs["CPU_Max_Freq_GHz"],
        "CPU_Cache_L3_MB":   specs["CPU_Cache_L3_MB"],
        "RAM_Type":          specs["RAM_Type"],
        "RAM_Freq":          specs["RAM_Freq"],
        "RAM_Size_GB":       specs["RAM_Size_GB"],
        "RAM_Slots":         specs["RAM_Slots"],
        "Nakopychuvach_SSD": specs["Nakopychuvach_SSD"],
        "Display_Diagonal":  specs["Display_Diagonal"],
        "Display_Max_Resolution": specs["Display_Max_Resolution"],
        "Display_Matrix_Type": specs["Display_Matrix_Type"],
        "Display_Cover":     specs["Display_Cover"],
        "Display_Brightness_nits": specs["Display_Brightness_nits"],
        "Display_Refresh_Rate": specs["Display_Refresh_Rate"],
        "Video_Brand":       specs["Video_Brand"],
        "GPU_Type":          specs["GPU_Type"],
        "GPU_Model":         specs["GPU_Model"],
        "GPU_Memory_MB":     specs["GPU_Memory_MB"],
        "OS":                specs["OS"],
        "Battery_Capacity_Wh": specs["Battery_Capacity_Wh"],
        "Camera_MP":         specs["Camera_MP"],
        "USB_TypeA":         specs["USB_TypeA"],
        "USB_TypeC":         specs["USB_TypeC"],
        "HDMI":              specs["HDMI"],
        "Keyboard_Waterproof":     specs["Keyboard_Waterproof"],
        "Keyboard_Ukrainian":      specs["Keyboard_Ukrainian"],
        "Keyboard_Pointing_Device": specs["Keyboard_Pointing_Device"],
        "Network_3G4G":      specs["Network_3G4G"],
        "Bluetooth":         specs["Bluetooth"],
        "WiFi":              specs["WiFi"],
        "LAN_Mbps":          specs["LAN_Mbps"],
        "Certificates":      specs["Certificates"],
        "Warranty":          specs["Warranty"],
        "Merezha_WiFi":      specs["Merezha_WiFi"],
        "Tsina_UAH":         price,
        "URL":               buy_url,
    }


def fetch_asus_shop_records(pg, shop_url, catalog_name, debug=False):
    """
    Fetch all SKU records from an ASUS /shop/ page.
    Returns list of record dicts (one per SKU / configuration).
    """
    try:
        # ASUS product comparison tables live on the /shop/ sub-page
        url_to_load = shop_url
        if '/shop' not in shop_url:
            url_to_load = shop_url.rstrip('/') + '/shop/'

        pg.goto(url_to_load, wait_until="load", timeout=DETAIL_PAGE_TIMEOUT)

        # Wait for comparison cards (up to 25 s)
        try:
            pg.wait_for_selector("[class*='compareCartItem']", timeout=25_000)
        except PWTimeout:
            # If still showing loading spinner, wait for it to clear
            try:
                pg.wait_for_selector("[class*='PageLoading']", state="hidden", timeout=20_000)
                pg.wait_for_timeout(2_000)
            except PWTimeout:
                pass

        html = pg.content()
        if debug:
            with open("debug_asus_detail.html", "w", encoding="utf-8") as f:
                f.write(html)
            print("    [debug] shop page HTML saved -> debug_asus_detail.html")

        soup = BeautifulSoup(html, "lxml")
        records = []

        # ── Per-SKU comparison cards (use cardInfo to exclude nested sub-elements)
        sku_cards = soup.select("[class*='compareCartItem'][class*='cardInfo']")
        if not sku_cards:
            sku_cards = soup.select("[class*='compareCartItem']")

        # ── Comparison table rows: list of (title, [value_per_sku]) ──────────
        _BUY_WORDS = {'купити', 'buy', 'придбати', 'замовити', 'add to cart'}
        compare_rows = []
        for box in soup.select("[class*='compareListBox']"):
            title_el = box.select_one("[class*='TitleMenuName']")
            value_els = box.select("[class*='compareListSpecItem']")
            if title_el and value_els:
                title = title_el.get_text(strip=True)
                values = [el.get_text(separator=" ", strip=True) for el in value_els]
                # Skip "Купити" button rows
                if all(v.strip().lower() in _BUY_WORDS or not v.strip() for v in values):
                    continue
                compare_rows.append((title, values))

        # ── pdKeySpec rows (key bullet specs per card) ────────────────────────
        key_spec_divs = soup.select("[class*='pdKeySpec']")

        num_skus = max(len(sku_cards), len(key_spec_divs),
                       max((len(v) for _, v in compare_rows), default=0))
        n_compare_cols = max((len(v) for _, v in compare_rows), default=0)
        print(f"      SKU cards={len(sku_cards)}, compare_rows={len(compare_rows)} (cols={n_compare_cols}), key_specs={len(key_spec_divs)} -> {num_skus} SKU(s) [{url_to_load.split('asus.com')[-1][:50]}]", flush=True)
        if debug and compare_rows:
            print("      [debug] First 5 compare_rows:")
            for title, vals in compare_rows[:5]:
                print(f"        '{title}' -> {vals[:3]}")
            print(f"      [debug] First sku_card texts: {[c.get_text(' ', strip=True)[:60] for c in sku_cards[:3]]}")

        if num_skus == 0:
            # Fallback: try broader selectors
            sku_cards = soup.select("[class*='productItem'], [class*='ProductItem'], [class*='pdCard']")
            compare_rows_alt = []
            for box in soup.select("[class*='specItem'], [class*='SpecItem'], [class*='specRow']"):
                title_el = box.select_one("[class*='specTitle'], [class*='specName'], [class*='title']")
                value_els = box.select("[class*='specValue'], [class*='value']")
                if title_el and value_els:
                    title = title_el.get_text(strip=True)
                    values = [el.get_text(separator=" ", strip=True) for el in value_els]
                    compare_rows_alt.append((title, values))
            if compare_rows_alt:
                compare_rows = compare_rows_alt
            num_skus = max(len(sku_cards), max((len(v) for _, v in compare_rows), default=0))
            if num_skus:
                print(f"      [fallback selectors] SKU cards={len(sku_cards)}, compare_rows={len(compare_rows)} -> {num_skus}", flush=True)
            else:
                slug = url_to_load.rstrip('/').split('/')[-2] if url_to_load.rstrip('/').endswith('/shop') else url_to_load.rstrip('/').split('/')[-1]
                debug_path = f"debug_asus_zero_{slug[:40]}.html"
                with open(debug_path, "w", encoding="utf-8") as f:
                    f.write(html)
                print(f"      [debug] 0-SKU page HTML saved -> {debug_path}", flush=True)

        _SKIP_LINES = {'менше', 'більше', 'more', 'less', 'докладніше', 'детальніше', 'купити', 'buy'}

        for i in range(num_skus):
            # Model number from card
            model_num = "-"
            price     = "-"
            buy_url   = url_to_load
            series_name = catalog_name

            if i < len(sku_cards):
                card = sku_cards[i]
                pn_el = card.select_one("[class*='pdModelName']")
                if pn_el:
                    model_num = pn_el.get_text(strip=True)
                name_el = card.select_one("[class*='pdName']")
                if name_el:
                    series_name = name_el.get_text(strip=True)
                price_el = card.select_one("[class*='pdPrice']")
                if price_el:
                    price = re.sub(r"[^\d\s]", "", price_el.get_text()).strip()
                buy_el = card.select_one(
                    "a[class*='pdBuyBtn'], a[class*='pdCard'], a[class*='pdSpec'], "
                    "a[class*='buyBtn'], a[class*='specBtn']"
                )
                if buy_el:
                    href = buy_el.get("href", "")
                    buy_url = href if href.startswith("http") else (BASE_URL + href if href.startswith("/") else url_to_load)

            # Ensure each SKU gets a unique URL so dedup doesn't collapse all to one row
            if buy_url == url_to_load:
                buy_url = url_to_load + (f"?sku={model_num}" if model_num != "-" else f"#sku{i}")

            # Build spec_items for this SKU
            spec_items = [series_name]
            if model_num != "-":
                spec_items.append(model_num)

            # Key specs: split by newline (bullets may be in spans inside a single <li>)
            if i < len(key_spec_divs):
                raw_text = key_spec_divs[i].get_text(separator='\n', strip=True)
                for line in raw_text.split('\n'):
                    line = line.strip()
                    if line and len(line) > 3 and line.lower() not in _SKIP_LINES:
                        spec_items.append(line)

            # Comparison table column i
            for title, values in compare_rows:
                if i < len(values) and values[i].strip():
                    spec_items.append(f"{title} {values[i]}")

            records.append(_build_record(series_name, model_num, price, buy_url, spec_items))

        return records

    except Exception as e:
        print(f"    [!] shop page error ({shop_url}): {e}")
        return []


# ── Catalog card extraction ───────────────────────────────────────────────────

# ASUS store uses several possible card selectors — try them all
CARD_SELECTORS = [
    "div.ProductItem_ProductItem__",
    "div[class*='ProductItem']",
    "div[class*='product-item']",
    "div[class*='product_item']",
    "article[class*='product']",
    "li[class*='product']",
    "div[class*='ProductCard']",
    "div[class*='product-card']",
    "[data-product-id]",
]

LINK_SELECTORS = [
    "a[class*='ProductItem']",
    "a[class*='product']",
    "a[href*='/laptops/']",
    "a[href*='/notebook']",
    "a",
]

PRICE_SELECTORS = [
    "[class*='price']",
    "[class*='Price']",
    "[class*='cost']",
    "[data-price]",
]

NAME_SELECTORS = [
    "[class*='title']",
    "[class*='Title']",
    "[class*='name']",
    "[class*='Name']",
    "h2", "h3", "h4",
]


def _extract_cards_generic(soup):
    """Generic ASUS card extractor using heuristic selectors."""
    records_raw = []

    for card_sel in CARD_SELECTORS:
        try:
            cards = soup.select(card_sel)
        except Exception:
            cards = []
        if not cards:
            continue

        for card in cards:
            # Find product link
            link_el = None
            for lsel in LINK_SELECTORS:
                link_el = card.select_one(lsel)
                if link_el:
                    break
            if not link_el:
                continue

            href = link_el.get("href", "")
            if not href or href == "#":
                continue
            url = BASE_URL + href if href.startswith("/") else href
            if "asus.com" not in url:
                continue
            # Only laptop/notebook product pages
            if not any(kw in url.lower() for kw in ['/laptops/', '/notebook']):
                continue

            # Find name
            name = ""
            for nsel in NAME_SELECTORS:
                name_el = card.select_one(nsel)
                if name_el:
                    name = name_el.get_text(strip=True)
                    if name:
                        break
            if not name:
                name = link_el.get_text(strip=True)
            if not name:
                continue

            # Find price
            price = "-"
            for psel in PRICE_SELECTORS:
                price_el = card.select_one(psel)
                if price_el:
                    price_text = price_el.get_text()
                    price = re.sub(r"[^\d\s]", "", price_text).strip()
                    if price:
                        break

            records_raw.append((name, url, price, card))

        if records_raw:
            break

    return records_raw



# Тексти кнопок/посилань, які НЕ є назвами товарів
_NAV_TEXTS = {
    "докладніше", "детальніше", "дізнатися більше", "вибрати магазин",
    "купити", "де купити", "read more", "learn more", "buy now",
    "compare", "порівняти", "add to cart", "wishlist",
}

def _is_product_name(text):
    """True if text looks like a real product name, not a nav button."""
    t = text.strip().lower()
    if not t or len(t) < 4:
        return False
    if t in _NAV_TEXTS:
        return False
    # Навігаційні тексти часто <= 2 слів та без цифр/брендових слів
    asus_brands = {"asus", "rog", "tuf", "zenbook", "vivobook", "expertbook",
                   "proart", "chromebook", "studiobook", "zephyrus", "flow"}
    has_brand = any(b in t for b in asus_brands)
    has_digit = bool(re.search(r'\d', t))
    # Потрібен або бренд, або цифра в назві (серійний номер / рік / діагональ)
    return has_brand or has_digit


def _extract_cards_js(pg):
    """Extract product data via JavaScript evaluation when HTML selectors fail."""
    try:
        data = pg.evaluate("""
            (() => {
                const NAV = new Set([
                    'докладніше','детальніше','дізнатися більше','вибрати магазин',
                    'купити','де купити','read more','learn more','buy now',
                    'compare','порівняти','add to cart','wishlist'
                ]);
                const results = [];
                // Collect links that point to individual product pages
                // ASUS URLs: /ua-ua/laptops/asus-vivobook-s16-s3607/  (slug ≥8 alphanumeric+dash chars)
                const links = Array.from(document.querySelectorAll('a[href]')).filter(a => {
                    const h = a.getAttribute('href') || '';
                    const t = a.textContent.trim().toLowerCase();
                    const isProductUrl = /\\/[A-Za-z0-9-]{8,}/.test(h) &&
                                         (h.includes('/laptops/') || h.includes('/Laptops/') || h.includes('/notebook')) &&
                                         !h.includes('#');
                    return isProductUrl && !NAV.has(t) && t.length > 3;
                });
                // Group by href, pick best (longest non-part-number) name
                const byHref = {};
                links.forEach(a => {
                    const href = a.getAttribute('href');
                    const name = a.textContent.trim();
                    if (NAV.has(name.toLowerCase())) return;
                    if (!byHref[href]) {
                        let price = '-';
                        const parent = a.closest('[class*="product"], [class*="Product"], li, article') || a.parentElement;
                        if (parent) {
                            const priceEl = parent.querySelector('[class*="price"], [class*="Price"]');
                            if (priceEl) price = priceEl.textContent.replace(/[^\\d\\s]/g, '').trim();
                        }
                        byHref[href] = { name, href, price };
                    } else {
                        // Prefer name that contains a space (not a bare part number)
                        const cur = byHref[href].name;
                        if (name.includes(' ') && !cur.includes(' ')) {
                            byHref[href].name = name;
                        } else if (name.length > cur.length && name.includes(' ')) {
                            byHref[href].name = name;
                        }
                    }
                });
                Object.values(byHref).forEach(item => results.push(item));
                return results;
            })()
        """)
        return data or []
    except Exception:
        return []


def parse_cards_from_page(pg, detail_pg=None, collected_so_far=0):
    """Parse product cards from the current page loaded in `pg`."""
    html = pg.content()
    soup = BeautifulSoup(html, "lxml")
    records = []

    raw_cards = _extract_cards_generic(soup)

    # Fallback: JS extraction
    if not raw_cards:
        js_data = _extract_cards_js(pg)
        for item in js_data:
            name  = item.get("name", "").strip()
            href  = item.get("href", "")
            price = item.get("price", "-")
            if not name or not href:
                continue
            url = BASE_URL + href if href.startswith("/") else href
            raw_cards.append((name, url, price, None))

    # Фільтрація: прибираємо навігаційні тексти та дублі по URL
    seen_urls = set()
    filtered = []
    for item in raw_cards:
        name, url, price, _card = item
        if not _is_product_name(name):
            continue
        if url in seen_urls:
            continue
        seen_urls.add(url)
        filtered.append(item)
    raw_cards = filtered

    for idx, (name, url, price, _card) in enumerate(raw_cards, 1):
        # Зупинка при досягненні MAX_ITEMS
        if collected_so_far + len(records) >= MAX_ITEMS:
            break
        print(f"    [{idx}/{len(raw_cards)}] {name[:60]}", flush=True)
        specs_items = [name]

        # Collect any visible spec items from card HTML
        if _card is not None:
            for li in _card.select("li, [class*='spec'], [class*='Spec'], [class*='feature']"):
                text = li.get_text(strip=True)
                if text and text != name and len(text) < 300:
                    specs_items.append(text)

        if detail_pg is not None and LOAD_DETAIL_PAGES:
            # ASUS shop pages contain multiple SKUs — expand into separate rows
            sku_records = fetch_asus_shop_records(
                detail_pg, url, catalog_name=name,
                debug=(idx == 1 and len(records) == 0)
            )
            if sku_records:
                print(f"      -> {len(sku_records)} SKU(s)", flush=True)
                records.extend(sku_records)
                continue  # skip fallback single-record path

        # Fallback: no detail page or shop page returned nothing
        records.append(_build_record(
            series_name=name, model_num="-", price=price,
            buy_url=url, spec_items=specs_items
        ))

    return records


# ── Page loading ──────────────────────────────────────────────────────────────

def _count_products(pg):
    """Return number of detectable product elements on page."""
    selectors = (
        "[class*='ProductItem'], [class*='product-item'], "
        "[class*='ProductCard'], [class*='product-card'], "
        "[data-product-id]"
    )
    try:
        return pg.evaluate(f"document.querySelectorAll(`{selectors}`).length")
    except Exception:
        return 0


def load_page_and_wait(pg, url):
    """Navigate to url, wait for product cards to appear. Returns True if found."""
    for attempt in range(2):
        try:
            pg.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
            break
        except Exception as err:
            if attempt >= 1:
                print(f"    Load error: {err}")
                return False
            time.sleep(1)

    # Dismiss cookie banner if present
    for sel in ["button[id*='accept'], button[class*='accept'], button[class*='cookie']"]:
        try:
            btn = pg.query_selector(sel)
            if btn:
                btn.click()
                time.sleep(0.3)
                break
        except Exception:
            pass

    # Wait for products to appear (up to 15 s)
    deadline = time.time() + 15
    last_count = 0
    while time.time() < deadline:
        count = _count_products(pg)
        if count > 0 and count == last_count:
            return True
        last_count = count
        time.sleep(0.4)

    # Last chance — also check for product links
    final_count = _count_products(pg)
    if final_count > 0:
        return True

    # Check product links as fallback
    try:
        link_count = pg.evaluate(
            "document.querySelectorAll('a[href*=\"/laptops/\"], a[href*=\"/notebook\"]').length"
        )
        if link_count > 3:
            print(f"    Found {link_count} product links (no card containers)")
            return True
    except Exception:
        pass

    # Save debug HTML
    html = pg.content()
    with open("debug_asus_page.html", "w", encoding="utf-8") as f:
        f.write(html)
    print("    No products found — debug HTML saved to debug_asus_page.html")
    return False


def detect_pagination(pg):
    """Detect total pages and pagination URL pattern for ASUS store."""
    # ASUS store typically uses ?page=N or /page/N or query params like ?p=N
    patterns_to_try = [
        CATALOG_URL + "?page={page}",
        CATALOG_URL + "?p={page}",
        CATALOG_URL + "page/{page}/",
    ]

    # Look for pagination elements to find total page count
    total_pages = 1
    try:
        # Try getting page count from pagination buttons
        nums = pg.evaluate("""
            (() => {
                const btns = Array.from(document.querySelectorAll(
                    '[class*="pagination"] a, [class*="pagination"] button, [class*="page-item"] a'
                ));
                const nums = btns.map(b => parseInt(b.textContent.trim(), 10)).filter(n => !isNaN(n) && n > 0);
                return nums.length ? Math.max(...nums) : 1;
            })()
        """)
        if nums and nums > 1:
            total_pages = nums
    except Exception:
        pass

    # If can't detect, also check "X results" text
    if total_pages == 1:
        try:
            count_text = pg.evaluate("""
                (() => {
                    const el = document.querySelector('[class*="result"], [class*="count"], [class*="total"]');
                    return el ? el.textContent : '';
                })()
            """)
            m = re.search(r'(\d+)', count_text or "")
            if m:
                items_total = int(m.group(1))
                # ASUS store usually shows ~24 items per page
                total_pages = max(1, (items_total + 23) // 24)
        except Exception:
            pass

    return total_pages, patterns_to_try[0]


# ── Excel formatting ──────────────────────────────────────────────────────────

def format_xlsx(path, total):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "ASUS Notebooks"

    UA = {
        "Seria": "Серія", "Model": "Модель", "Part Number": "Part Number",
        "CPU_Brand": "Бренд CPU", "CPU_Model": "Модель CPU", "CPU_Cores": "Кількість ядер",
        "CPU_Threads": "Кількість потоків", "CPU_Base_Freq_GHz": "Номінальна частота, ГГц",
        "CPU_Max_Freq_GHz": "Максимальна частота, ГГц", "CPU_Cache_L3_MB": "Кеш L3, МБ",
        "RAM_Type": "Тип ОЗП", "RAM_Freq": "Частота ОЗП",
        "RAM_Size_GB": "Обсяг ОЗП, ГБ", "RAM_Slots": "Кількість слотів",
        "Nakopychuvach_SSD": "M.2 SSD, ГБ",
        "Display_Diagonal": "Діагональ екрана", "Display_Max_Resolution": "Макс. роздільна здатність",
        "Display_Matrix_Type": "Тип матриці", "Display_Cover": "Покриття екрану",
        "Display_Brightness_nits": "Яскравість, ніт", "Display_Refresh_Rate": "Частота оновлення",
        "Video_Brand": "Видео_Бренд", "GPU_Type": "Тип_GPU", "GPU_Model": "Модель GPU",
        "GPU_Memory_MB": "Обсяг GPU, МБ", "OS": "ОС",
        "Battery_Capacity_Wh": "Енергетична ємність, Вт*год",
        "Camera_MP": "WEB-камера, Мп", "USB_TypeA": "USB Type-A", "USB_TypeC": "USB Type-C",
        "HDMI": "HDMI, шт.",
        "Keyboard_Waterproof": "Клавіатура (вологозахист)", "Keyboard_Ukrainian": "Українська мова",
        "Keyboard_Pointing_Device": "Маніпулятори", "Network_3G4G": "3G/4G",
        "Bluetooth": "Bluetooth", "WiFi": "Wi-Fi", "LAN_Mbps": "LAN RJ-45, Мбіт/с",
        "Certificates": "Сертифікати", "Warranty": "Термін базової гарантії від виробника",
        "Merezha_WiFi": "Мережа / Wi-Fi", "Tsina_UAH": "Ціна (UAH)", "URL": "URL",
    }
    for i, k in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=i).value = UA.get(k, k)

    # ASUS brand colour: dark navy
    HFILL = PatternFill("solid", fgColor="00539B")
    AFILL = PatternFill("solid", fgColor="EEF4FF")
    t   = Side(style="thin", color="BBBBBB")
    brd = Border(left=t, right=t, top=t, bottom=t)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
    ws.row_dimensions[1].height = 32

    for r in range(2, ws.max_row + 1):
        for cell in ws[r]:
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = AFILL if r % 2 == 0 else PatternFill()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = brd
        # No explicit height -> Excel auto-fits based on cell content + wrap_text

    for i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    url_col = len(COLUMNS)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=url_col)
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    info = wb.create_sheet("Info")
    info["A1"], info["B1"] = "Сформовано:",  datetime.now().strftime("%d.%m.%Y %H:%M")
    info["A2"], info["B2"] = "Джерело:",     CATALOG_URL
    info["A3"], info["B3"] = "Моделей:",     total
    for c in ["A1", "A2", "A3"]:
        info[c].font = Font(bold=True, name="Arial")
    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  ASUS Notebooks Parser — asus.com/ua-ua/store/laptops/")
    print("=" * 60)

    all_records = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
        )
        ctx = browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="uk-UA",
        )
        pg = ctx.new_page()

        # Page 1
        print("\nЗавантаження сторінки 1...")
        ok = load_page_and_wait(pg, CATALOG_URL)
        if not ok:
            print("ПОМИЛКА: товари не знайдено. Зупинка.")
            browser.close()
            return

        total_pages, url_pattern = detect_pagination(pg)
        print(f"Сторінок: {total_pages}  (приблизно {total_pages * 24} товарів)")
        print()

        # Parse page 1 (already loaded)
        recs = parse_cards_from_page(pg, pg, collected_so_far=0)
        all_records.extend(recs)
        print(f"  [1/{total_pages}] зібрано {len(recs)} (всього: {len(all_records)})")

        # Pages 2..N
        for page_num in range(2, total_pages + 1):
            if len(all_records) >= MAX_ITEMS:
                print(f"\nДосягнуто ліміту {MAX_ITEMS} товарів.")
                break

            print(f"  [{page_num}/{total_pages}] ...", end=" ", flush=True)

            success  = False
            retry    = 0

            while not success and retry < 2:
                url = url_pattern.replace("{page}", str(page_num))
                ok  = load_page_and_wait(pg, url)
                if ok:
                    recs = parse_cards_from_page(pg, pg, collected_so_far=len(all_records))
                    if recs:
                        all_records.extend(recs)
                        print(f"зібрано {len(recs)} (всього: {len(all_records)})")
                        success = True
                        break

                retry += 1
                if retry < 2:
                    print("\n  [!] Повтор...", end=" ", flush=True)
                    time.sleep(2)

            if not success:
                # Try scroll-based pagination: click "Завантажити ще" / "Load more"
                try:
                    more_btn = pg.query_selector(
                        "button[class*='load-more'], button[class*='LoadMore'], "
                        "button[class*='show-more'], a[class*='load-more']"
                    )
                    if more_btn:
                        more_btn.click()
                        time.sleep(2)
                        recs = parse_cards_from_page(pg, pg, collected_so_far=len(all_records))
                        if recs:
                            all_records.extend(recs)
                            print(f"зібрано {len(recs)} (всього: {len(all_records)})")
                            success = True
                except Exception:
                    pass

            if not success:
                print("0 — пропуск")

        browser.close()

    # Save
    print()
    all_records = all_records[:MAX_ITEMS]
    print(f"Збереження Excel ({len(all_records)} моделей)...")
    df = pd.DataFrame(all_records, columns=COLUMNS)
    before = len(df)
    valid_mask  = df["Part Number"] != "-"
    valid_df    = df[valid_mask].drop_duplicates(subset=["Part Number"], keep="first")
    # Also deduplicate by URL for items without Part Number
    invalid_df  = df[~valid_mask].drop_duplicates(subset=["URL"], keep="first")
    df = pd.concat([valid_df, invalid_df], ignore_index=True)
    if len(df) < before:
        print(f"Дублікатів виключено: {before - len(df)}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False, sheet_name="ASUS Notebooks")
    print("Форматування...")
    format_xlsx(OUTPUT_PATH, len(df))

    print()
    print("=" * 60)
    print(f"ГОТОВО!  {OUTPUT_PATH}")
    print(f"Моделей у таблиці: {len(df)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
